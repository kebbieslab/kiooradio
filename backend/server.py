from fastapi import FastAPI, APIRouter, HTTPException, Response, Form, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import JSONResponse, HTMLResponse
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
import os
import json
import secrets
import httpx
import logging
from datetime import datetime, time, date, timezone, timedelta
from enum import Enum

# New imports for document processing
import fitz  # PyMuPDF
import io
from PIL import Image
import aiohttp
import aiofiles
import hashlib
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pandas as pd
import csv
from io import StringIO
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create directories for document processing
TEMP_DIR = ROOT_DIR / "temp"
THUMBNAILS_DIR = ROOT_DIR / "static" / "thumbnails"
CACHE_DIR = ROOT_DIR / "cache"

# Create directories if they don't exist
for directory in [TEMP_DIR, THUMBNAILS_DIR, CACHE_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

# MongoDB connection with error handling
try:
    mongo_url = os.environ.get('MONGO_URL')
    if not mongo_url:
        raise ValueError("MONGO_URL environment variable is required")
    
    db_name = os.environ.get('DB_NAME')
    if not db_name:
        raise ValueError("DB_NAME environment variable is required")
    logger.info(f"Connecting to MongoDB at {mongo_url[:20]}... using database: {db_name}")
    
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]
    
except Exception as e:
    logger.error(f"Failed to initialize MongoDB connection: {e}")
    raise

# Create the main app without a prefix
app = FastAPI(
    title="Kioo Radio API", 
    version="1.0.0",
    docs_url="/docs" if os.environ.get("ENVIRONMENT") != "production" else None,
    redoc_url="/redoc" if os.environ.get("ENVIRONMENT") != "production" else None
)

# Configure CORS
cors_origins = os.environ.get('CORS_ORIGINS', '*')
if cors_origins == '*':
    allowed_origins = ["*"]
else:
    allowed_origins = [origin.strip() for origin in cors_origins.split(',')]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint for deployment"""
    try:
        # Test database connection
        await client.admin.command('ping')
        return {
            "status": "healthy",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "database": "connected"
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(status_code=503, detail="Service unavailable")

# Serve static files for thumbnails
from fastapi.staticfiles import StaticFiles
app.mount("/api/static", StaticFiles(directory=ROOT_DIR / "static"), name="static")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Enums
class ProgramLanguage(str, Enum):
    ENGLISH = "english"
    FRENCH = "french"
    KISSI = "kissi"
    KRIO = "krio"
    MIXED = "mixed"
    MANDINGO = "mandingo"
    FULA = "fula"
    GBANDI = "gbandi"

class ProgramCategory(str, Enum):
    NEWS = "news"
    TALK = "talk"
    MUSIC = "music"
    EDUCATIONAL = "educational"
    YOUTH = "youth"
    FARMING = "farming"
    COMMUNITY = "community"
    BIBLE_TEACHING = "bible_teaching"
    INTERACTIVE = "interactive"
    DEVOTIONAL = "devotional"
    OUTREACH = "outreach"
    WORSHIP = "worship"
    LIVE_SERVICE = "live_service"
    SATELLITE = "satellite"
    SERMON = "sermon"
    TALK_SHOW = "talk_show"
    CHRISTIAN_TEACHING = "christian_teaching"

class DayOfWeek(str, Enum):
    MONDAY = "monday"
    TUESDAY = "tuesday"
    WEDNESDAY = "wednesday"
    THURSDAY = "thursday"
    FRIDAY = "friday"
    SATURDAY = "saturday"
    SUNDAY = "sunday"

# Models
class Program(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: str
    host: str
    language: ProgramLanguage
    category: ProgramCategory
    day_of_week: DayOfWeek
    start_time: str  # Format: "HH:MM"
    duration_minutes: int
    is_live: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ProgramCreate(BaseModel):
    title: str
    description: str
    host: str
    language: ProgramLanguage
    category: ProgramCategory
    day_of_week: DayOfWeek
    start_time: str
    duration_minutes: int

class ImpactStory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    content: str
    author_name: str
    author_location: str
    image_url: Optional[str] = None
    is_featured: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ImpactStoryCreate(BaseModel):
    title: str
    content: str
    author_name: str
    author_location: str
    image_url: Optional[str] = None
    is_featured: bool = False

class NewsUpdate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    content: str
    excerpt: str
    author: str
    image_url: Optional[str] = None
    is_published: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)

class NewsUpdateCreate(BaseModel):
    title: str
    content: str
    excerpt: str
    author: str
    image_url: Optional[str] = None
    is_published: bool = True

class Donation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    donor_name: str
    donor_email: str
    amount: float
    currency: str = "USD"
    donation_type: str  # "one-time" or "monthly"
    message: Optional[str] = None
    is_anonymous: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)

class DonationCreate(BaseModel):
    donor_name: str
    donor_email: str
    amount: float
    currency: str = "USD"
    donation_type: str
    message: Optional[str] = None
    is_anonymous: bool = False

class StationSettings(BaseModel):
    stationWhatsAppNumber: str = "+231778383703"
    stationWhatsAppDigitsOnly: str = "231778383703" 
    stationEmail: str = "info@kiooradio.org"
    stationName: str = "Kioo Radio 98.1 FM"

class ContactMessage(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    phone: Optional[str] = None
    subject: str = ""
    message: str
    partnerRef: Optional[str] = None
    pastorName: Optional[str] = None
    churchName: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ContactMessageCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    subject: str = ""
    message: str
    partnerRef: Optional[str] = None
    pastorName: Optional[str] = None
    churchName: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None

class RadioStatus(BaseModel):
    is_live: bool
    current_program: Optional[str] = None
    next_program: Optional[str] = None
    listener_count: int = 0

class MajorGiftPledge(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    phone: Optional[str] = None
    amount: float
    designation: str  # Solar, Studios, Programming, "Where most needed"
    pledgeDate: str
    message: Optional[str] = None
    status: str = "new"  # new, contacted, fulfilled
    created_at: datetime = Field(default_factory=datetime.utcnow)

class MajorGiftPledgeCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    amount: float
    designation: str
    pledgeDate: str
    message: Optional[str] = None

class MediaVideo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    youtubeUrl: str
    language: str = "en"  # en/fr
    description: Optional[str] = None
    publishedAt: Optional[datetime] = None
    order: int = 1
    featured: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class MediaPhoto(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    imageUrl: str
    alt: str
    caption: Optional[str] = None
    credit: Optional[str] = None
    takenAt: Optional[datetime] = None
    location: Optional[str] = None  # county/city
    language: Optional[str] = "en"  # en/fr/optional
    featured: bool = False
    order: int = 1
    consentToPublish: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class MediaSettings(BaseModel):
    mediaPromoEnabled: bool = True
    videoEmbedPrivacy: str = "youtube-nocookie"
    lightbox: str = "photoswipe"
    photoPlaceholderStrategy: str = "lqip-blur"
    maxColumnsDesktop: int = 3

class MajorGiftsSettings(BaseModel):
    calUrl: Optional[str] = ""
    caseForSupportUrl: Optional[str] = ""
    impactBriefUrl: Optional[str] = ""
    budgetOverviewUrl: Optional[str] = ""
    namingOpportunities: List[str] = [
        "Studio Naming — $25,000+",
        "Transmitter Room — $15,000+", 
        "Solar Array — $35,000+",
        "Youth Programming — $10,000+",
        "Cross-Border Outreach — $20,000+"
    ]

class PaymentSettings(BaseModel):
    directLiberiaEmbedType: str = "iframe"
    directLiberiaEmbedCode: str = ""
    directLiberiaFallbackUrl: str = ""
    successRedirect: str = "/donate/thank-you"
    cancelRedirect: str = "/donate"

class AboutPageSettings(BaseModel):
    # Vision 2005 section
    visionTitle: str = "The Vision (2005)"
    visionContent: str = """In 2005, while studying at Media Village, a media school of YWMA in Cape Town, South Africa, God placed a burden on Joseph Kebbie's heart. The vision was clear: "Return to Liberia and start a radio station to reach my people." Originally, God wanted us to first start the Kissi radio station, but in His own ways, He led us to start Vox Radio first in 2017. Vox Radio began in a shipping container and has grown to serve over 3.2 million people across 8 counties in Liberia with grant support from Elmer H. Schmidt Christian Broadcasting Fund provided by Community Foundation of the North State. Now, Kioo Radio emerges as the fulfillment of God's original vision—a beacon of hope specifically for the Makona River Region, where the borders of Liberia, Sierra Leone, and Guinea meet."""
    
    # Timeline section (reordered 2024 entries)
    timelineTitle: str = "From Vision to Launch"
    timelineItems: List[dict] = [
        {"year": "2005", "event": "Vision received in Cape Town: Start a radio station in Foya, Liberia to reach Sierra Leone and Guinea"},
        {"year": "2017", "event": "Vox Radio meant to be the second station was established in a shipping container, beginning with local community programming"},
        {"year": "2024", "event": "Vox Radio expands coverage to reach 3.2M+ people across 8 counties with grant support from Elmer H. Schmidt Christian Broadcasting Fund provided by Community Foundation of the North State"},
        {"year": "2024", "event": "Daniel Hatfield challenged me to trust God to begin Kioo Radio"},
        {"year": "2025", "event": "Broadcasting license approved for Kioo Radio 98.1FM"},
        {"year": "2025", "event": "Studio construction and equipment installation begins in Foya"},
        {"year": "2025", "event": "Kioo Radio 98.1FM official launch scheduled for November 13, 2025"}
    ]
    
    # Kissi people section
    kissiTitle: str = "Who Are the Kissi? (Our Heartland)"
    kissiContent: str = """The Kissi people are a proud West African ethnic group primarily inhabiting the mountainous regions where Liberia, Sierra Leone, and Guinea converge. Known for their rich cultural heritage, agricultural expertise, and strong family values, the Kissi have maintained their traditions while adapting to modern challenges. Their language, also called Kissi, serves as a unifying force across national borders. Kioo Radio is named with deep respect for this community—"Kissi" meaning "Gift" in their native language—reflecting our mission to be a gift of God's love and truth to this region. The station serves not only the Kissi people but all communities in this tri-border area, broadcasting in multiple languages to reach every heart with the Gospel message. Originally, God's vision was to start with the Kissi radio station, but He led us to establish Vox Radio first in 2017, preparing us for this moment when Kioo Radio would fulfill His original calling."""
    
    # Document URLs
    radioProjectPptUrl: str = "https://customer-assets.emergentagent.com/job_radio-geo-detect/artifacts/aja36zyu_Radio%20Project11.ppt"
    maruRadioProposalPdfUrl: str = "https://customer-assets.emergentagent.com/job_radio-geo-detect/artifacts/r6bt039z_maru_radio_proposal.PDF"
    
    # Preview image URLs (generated dynamically)
    radioProjectPreviewImages: List[str] = []
    maruRadioProposalPreviewImages: List[str] = []

# Document processing functions
async def download_file_from_url(url: str, timeout: int = 30) -> bytes:
    """Download file from URL with proper error handling."""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=timeout)) as response:
                if response.status == 200:
                    return await response.read()
                else:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Failed to download file from {url}: HTTP {response.status}"
                    )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")

def generate_pdf_thumbnails(pdf_bytes: bytes, max_pages: int = 3) -> List[bytes]:
    """Generate thumbnail images from PDF pages."""
    thumbnails = []
    
    try:
        # Open PDF from bytes
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        # Set matrix for high quality output (300 DPI)
        mat = fitz.Matrix(2.0, 2.0)  # 2x scale for better quality
        
        page_count = min(len(doc), max_pages)
        
        for page_num in range(page_count):
            page = doc[page_num]
            
            # Create pixmap with high quality
            pix = page.get_pixmap(matrix=mat, alpha=False)
            
            # Convert to PIL Image for thumbnail processing
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            
            # Create thumbnail while maintaining aspect ratio
            img.thumbnail((400, 300), Image.Resampling.LANCZOS)
            
            # Convert back to bytes
            thumbnail_bytes = io.BytesIO()
            img.save(thumbnail_bytes, format='PNG', optimize=True, quality=90)
            thumbnails.append(thumbnail_bytes.getvalue())
        
        doc.close()
        
    except Exception as e:
        print(f"Error generating PDF thumbnails: {e}")
        
    return thumbnails

async def generate_document_previews():
    """Generate preview images for both documents."""
    try:
        # Generate cache keys
        ppt_cache_key = hashlib.md5("radio_project_ppt".encode()).hexdigest()
        pdf_cache_key = hashlib.md5("maru_radio_proposal_pdf".encode()).hexdigest()
        
        preview_urls = {
            "radioProjectPreviewImages": [],
            "maruRadioProposalPreviewImages": []
        }
        
        # Process PDF document
        try:
            pdf_url = "https://customer-assets.emergentagent.com/job_radio-geo-detect/artifacts/r6bt039z_maru_radio_proposal.PDF"
            pdf_bytes = await download_file_from_url(pdf_url)
            pdf_thumbnails = generate_pdf_thumbnails(pdf_bytes, max_pages=3)
            
            for i, thumbnail_data in enumerate(pdf_thumbnails):
                filename = f"pdf_{pdf_cache_key}_page_{i+1}.png"
                filepath = THUMBNAILS_DIR / filename
                
                # Save thumbnail
                async with aiofiles.open(filepath, 'wb') as f:
                    await f.write(thumbnail_data)
                
                preview_urls["maruRadioProposalPreviewImages"].append(f"/api/static/thumbnails/{filename}")
                
        except Exception as e:
            print(f"Error processing PDF: {e}")
        
        # For PowerPoint, we'll create a placeholder preview for now
        # since Aspose.Slides requires a license
        try:
            # Create placeholder preview images for PowerPoint
            for i in range(3):
                placeholder_img = Image.new('RGB', (400, 300), color=(240, 240, 240))
                # Add text to placeholder
                filename = f"ppt_{ppt_cache_key}_slide_{i+1}.png"
                filepath = THUMBNAILS_DIR / filename
                
                placeholder_img.save(filepath, format='PNG', optimize=True)
                preview_urls["radioProjectPreviewImages"].append(f"/api/static/thumbnails/{filename}")
                
        except Exception as e:
            print(f"Error creating PowerPoint placeholders: {e}")
        
        return preview_urls
        
    except Exception as e:
        print(f"Error in generate_document_previews: {e}")
        return {
            "radioProjectPreviewImages": [],
            "maruRadioProposalPreviewImages": []
        }
    wire: dict = {
        "beneficiaryName": "VOX Liberia",
        "beneficiaryAddress": "C/O Galcom International, P.O. Box 1211, Foya, Lofa County, Liberia",
        "bankName": "Ecobank Liberia Limited",
        "bankAddress": "Broad Street, Monrovia, Liberia",
        "accountNumber": "2040034567890123",
        "swift": "ECOCLRLR",
        "routingNumber": "011000028",
        "intermediaryBank": "Citibank N.A. New York",
        "intermediaryAddress": "399 Park Avenue, New York, NY 10043, USA",
        "intermediarySwift": "CITIUS33",
        "intermediaryAccount": "36083522",
        "wireInstructions": "For further credit to VOX Liberia Account #2040034567890123",
        "referenceNote": "Kioo Radio 98.1 FM Major Gift",
        "contactEmail": "info@kiooradio.org",
        "contactPhone": "+231 77 838 3703"
    }

class NewsletterSignup(BaseModel):
    email: str
    adminEmail: str

class NewsletterSignupCreate(BaseModel):
    email: str
    adminEmail: str

class ChurchPartner(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    pastorName: str
    churchName: str
    country: str  # Liberia, Sierra Leone, Guinea
    city: str
    altCityNames: List[str] = []
    onAirDaysTimes: Optional[str] = None
    contactPhone: Optional[str] = None
    whatsAppNumber: Optional[str] = None
    consentToDisplayContact: bool = False
    notes: Optional[str] = None
    photoUrl: Optional[str] = None
    isPublished: bool = True
    sortOrder: Optional[int] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ChurchPartnerCreate(BaseModel):
    pastorName: str
    churchName: str
    country: str
    city: str
    altCityNames: List[str] = []
    onAirDaysTimes: Optional[str] = None
    contactPhone: Optional[str] = None
    whatsAppNumber: Optional[str] = None
    consentToDisplayContact: bool = False
    notes: Optional[str] = None
    photoUrl: Optional[str] = None
    isPublished: bool = True
    sortOrder: Optional[int] = None

# Email Models
class ContactFormSubmission(BaseModel):
    name: str
    email: EmailStr
    subject: str
    message: str
    timestamp: Optional[datetime] = None

class NewsletterSubscription(BaseModel):
    email: EmailStr
    timestamp: Optional[datetime] = None

class VisitorAnalytics(BaseModel):
    ip_address: str
    country: Optional[str] = None
    city: Optional[str] = None
    page_url: str
    user_agent: str
    referrer: Optional[str] = None
    timestamp: Optional[datetime] = None

# Authentication for visitors page
security = HTTPBasic()

def authenticate_admin(credentials: HTTPBasicCredentials = Depends(security)):
    """Simple admin authentication for visitors page"""
    correct_username = secrets.compare_digest(credentials.username, "admin")
    correct_password = secrets.compare_digest(credentials.password, "kioo2025!")
    
    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username

# Visitor Analytics and Email Helper Functions
async def send_simple_email(to_email: str, subject: str, body: str):
    """Send email using a simple SMTP approach"""
    try:
        # For now, we'll just log the email - in production you'd use a real SMTP service
        logging.info(f"EMAIL SENT TO: {to_email}")
        logging.info(f"SUBJECT: {subject}")
        logging.info(f"BODY: {body}")
        return True
    except Exception as e:
        logging.error(f"Failed to send email: {e}")
        return False

async def get_ip_location(ip_address: str):
    """Get location information from IP address using ipapi.co"""
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(f"https://ipapi.co/{ip_address}/json/")
            if response.status_code == 200:
                data = response.json()
                return {
                    "country": data.get("country_name"),
                    "city": data.get("city"),
                    "region": data.get("region"),
                    "latitude": data.get("latitude"),
                    "longitude": data.get("longitude")
                }
    except Exception as e:
        logging.error(f"Failed to get IP location: {e}")
    
    return {"country": "Unknown", "city": "Unknown"}

# Visitor tracking endpoints
@api_router.post("/track-visitor")
async def track_visitor(visitor_data: dict):
    """Track visitor page visits with IP geolocation"""
    try:
        # Get client IP address (in production, handle proxies properly)
        client_ip = visitor_data.get("client_ip", "127.0.0.1")
        
        # Get location data
        location_data = await get_ip_location(client_ip)
        
        # Create visitor record
        visitor_record = VisitorAnalytics(
            ip_address=client_ip,
            country=location_data.get("country"),
            city=location_data.get("city"),
            page_url=visitor_data.get("page_url", ""),
            user_agent=visitor_data.get("user_agent", ""),
            referrer=visitor_data.get("referrer"),
            timestamp=datetime.now(timezone.utc)
        )
        
        # Store in database
        await db.visitor_analytics.insert_one(visitor_record.dict())
        
        return {"status": "success", "message": "Visitor tracked successfully"}
    except Exception as e:
        logging.error(f"Failed to track visitor: {e}")
        return {"status": "error", "message": "Failed to track visitor"}

@api_router.post("/track-click")
async def track_click(click_data: dict):
    """Track click events"""
    try:
        # Create click record
        click_record = {
            "id": str(uuid.uuid4()),
            "element_type": click_data.get("element_type"),
            "element_id": click_data.get("element_id"),
            "element_class": click_data.get("element_class"),
            "element_text": click_data.get("element_text"),
            "page_url": click_data.get("page_url"),
            "click_position": click_data.get("click_position"),
            "timestamp": datetime.now(timezone.utc)
        }
        
        # Store in database
        await db.click_analytics.insert_one(click_record)
        
        return {"status": "success", "message": "Click tracked successfully"}
    except Exception as e:
        logging.error(f"Failed to track click: {e}")
        return {"status": "error", "message": "Failed to track click"}

# Visitor analytics dashboard endpoints (protected)
@api_router.get("/visitors/management-stats")
async def get_visitor_stats(credentials: HTTPBasicCredentials = Depends(authenticate_admin)):
    """Get visitor statistics for the analytics dashboard"""
    try:
        # Get total visitors
        total_visitors = await db.visitor_analytics.count_documents({})
        
        # Get unique visitors (by IP)
        unique_visitors = len(await db.visitor_analytics.distinct("ip_address"))
        
        # Get visitors today
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        visitors_today = await db.visitor_analytics.count_documents({
            "timestamp": {"$gte": today}
        })
        
        # Get top countries
        top_countries_pipeline = [
            {"$group": {"_id": "$country", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        top_countries = await db.visitor_analytics.aggregate(top_countries_pipeline).to_list(10)
        
        # Get top pages
        top_pages_pipeline = [
            {"$group": {"_id": "$page_url", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        top_pages = await db.visitor_analytics.aggregate(top_pages_pipeline).to_list(10)
        
        # Get hourly traffic (last 24 hours)
        hourly_traffic = []
        for i in range(24):
            hour_start = datetime.now(timezone.utc) - timedelta(hours=i+1)
            hour_end = datetime.now(timezone.utc) - timedelta(hours=i)
            
            count = await db.visitor_analytics.count_documents({
                "timestamp": {"$gte": hour_start, "$lt": hour_end}
            })
            
            hourly_traffic.append({
                "hour": hour_start.strftime("%H:00"),
                "count": count
            })
        
        hourly_traffic.reverse()  # Show oldest to newest
        
        return {
            "total_visitors": total_visitors,
            "unique_visitors": unique_visitors,
            "visitors_today": visitors_today,
            "top_countries": top_countries,
            "top_pages": top_pages,
            "hourly_traffic": hourly_traffic
        }
    except Exception as e:
        logging.error(f"Failed to get visitor stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get visitor statistics")

@api_router.get("/visitors/recent")
async def get_recent_visitors(credentials: HTTPBasicCredentials = Depends(authenticate_admin)):
    """Get recent visitor activity"""
    try:
        recent_visitors = await db.visitor_analytics.find().sort("timestamp", -1).limit(50).to_list(50)
        
        # Convert datetime objects to strings for JSON serialization
        for visitor in recent_visitors:
            if '_id' in visitor:
                del visitor['_id']
            if 'timestamp' in visitor and isinstance(visitor['timestamp'], datetime):
                visitor['timestamp'] = visitor['timestamp'].isoformat()
        
        return recent_visitors
    except Exception as e:
        logging.error(f"Failed to get recent visitors: {e}")
        raise HTTPException(status_code=500, detail="Failed to get recent visitors")

@api_router.get("/visitors/clicks")
async def get_click_analytics(credentials: HTTPBasicCredentials = Depends(authenticate_admin)):
    """Get click analytics data"""
    try:
        # Get most clicked elements
        click_pipeline = [
            {"$group": {"_id": "$element_type", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        click_stats = await db.click_analytics.aggregate(click_pipeline).to_list(10)
        
        # Get recent clicks
        recent_clicks = await db.click_analytics.find().sort("timestamp", -1).limit(20).to_list(20)
        
        # Convert datetime objects to strings
        for click in recent_clicks:
            if '_id' in click:
                del click['_id']
            if 'timestamp' in click and isinstance(click['timestamp'], datetime):
                click['timestamp'] = click['timestamp'].isoformat()
        
        return {
            "click_stats": click_stats,
            "recent_clicks": recent_clicks
        }
    except Exception as e:
        logging.error(f"Failed to get click analytics: {e}")
        raise HTTPException(status_code=500, detail="Failed to get click analytics")

# Newsletter subscription endpoint
@api_router.post("/subscribe")
async def subscribe_newsletter(email_data: dict):
    """Handle newsletter subscription"""
    try:
        email = email_data.get("email")
        if not email:
            raise HTTPException(status_code=400, detail="Email is required")
        
        # Store subscription
        subscription_record = NewsletterSubscription(
            email=email,
            timestamp=datetime.now(timezone.utc)
        )
        
        await db.newsletter_subscriptions.insert_one(subscription_record.dict())
        
        # Send notification email to admin
        subject = f"New Newsletter Subscription - Kioo Radio"
        content = f"""
New newsletter subscription received:

Email: {email}
Timestamp: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC

---
Kioo Radio 98.1 FM
Broadcasting Faith, Hope and Love in Christ across the Makona River Region
        """
        
        await send_simple_email("admin@proudlyliberian.com", subject, content)
        
        return {"status": "success", "message": "Successfully subscribed to newsletter"}
    except Exception as e:
        logging.error(f"Failed to handle newsletter subscription: {e}")
        raise HTTPException(status_code=500, detail="Failed to subscribe to newsletter")

# Contact form submission endpoint  
@api_router.post("/contact-form")
async def submit_contact_form(form_data: dict):
    """Handle contact form submissions"""
    try:
        # Store contact form submission
        contact_record = ContactFormSubmission(
            name=form_data.get("name", ""),
            email=form_data.get("email", ""),
            subject=form_data.get("subject", ""),
            message=form_data.get("message", ""),
            timestamp=datetime.now(timezone.utc)
        )
        
        await db.contact_form_submissions.insert_one(contact_record.dict())
        
        # Send notification email to admin
        subject = f"New Contact Form Submission - {form_data.get('subject', 'No Subject')}"
        content = f"""
New contact form submission received:

Name: {form_data.get('name', 'Not provided')}
Email: {form_data.get('email', 'Not provided')}
Subject: {form_data.get('subject', 'No subject')}

Message:
{form_data.get('message', 'No message')}

Timestamp: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC

---
Kioo Radio 98.1 FM
Broadcasting Faith, Hope and Love in Christ across the Makona River Region
        """
        
        await send_simple_email("admin@proudlyliberian.com", subject, content)
        
        return {"status": "success", "message": "Contact form submitted successfully"}
    except Exception as e:
        logging.error(f"Failed to handle contact form submission: {e}")
        raise HTTPException(status_code=500, detail="Failed to submit contact form")

# Routes
@api_router.get("/")
async def root():
    return {"message": "Kioo Radio API - The Gift of Good News", "version": "1.0.0"}

@api_router.get("/radio/status", response_model=RadioStatus)
async def get_radio_status():
    # Placeholder for now - in real implementation, this would check actual streaming status
    return RadioStatus(
        is_live=True,
        current_program="Good Morning Kioo",
        next_program="Community Voice",
        listener_count=247
    )

# Programs endpoints
@api_router.post("/programs", response_model=Program)
async def create_program(program: ProgramCreate):
    program_dict = program.dict()
    program_obj = Program(**program_dict)
    await db.programs.insert_one(program_obj.dict())
    return program_obj

@api_router.get("/programs", response_model=List[Program])
async def get_programs(language: Optional[ProgramLanguage] = None, day: Optional[DayOfWeek] = None):
    filter_dict = {}
    if language:
        filter_dict["language"] = language
    if day:
        filter_dict["day_of_week"] = day
    
    programs = await db.programs.find(filter_dict).sort("start_time", 1).to_list(1000)
    
    # Convert database programs to API format, handling data inconsistencies
    converted_programs = []
    for program in programs:
        try:
            # Remove MongoDB _id field
            if '_id' in program:
                del program['_id']
            
            # Handle duration field mapping
            if 'duration' in program and 'duration_minutes' not in program:
                program['duration_minutes'] = program['duration']
            elif 'duration_minutes' not in program:
                program['duration_minutes'] = 30  # Default duration
            
            # Remove extra fields that aren't in the Program model
            extra_fields = ['duration', 'end_time', 'is_recurring', 'new_program', 'updated_at']
            for field in extra_fields:
                if field in program:
                    del program[field]
            
            # Convert datetime strings to datetime objects if needed
            if 'created_at' in program and isinstance(program['created_at'], str):
                from datetime import datetime
                program['created_at'] = datetime.fromisoformat(program['created_at'].replace('Z', '+00:00'))
            
            converted_programs.append(Program(**program))
        except Exception as e:
            print(f"Error converting program {program.get('id', 'unknown')}: {e}")
            continue
    
    return converted_programs

@api_router.get("/programs/schedule")
async def get_schedule():
    programs = await db.programs.find().sort([("day_of_week", 1), ("start_time", 1)]).to_list(1000)
    
    # Group by day of week
    schedule = {}
    for program in programs:
        try:
            # Remove MongoDB _id field
            if '_id' in program:
                del program['_id']
            
            # Handle duration field mapping
            if 'duration' in program and 'duration_minutes' not in program:
                program['duration_minutes'] = program['duration']
            elif 'duration_minutes' not in program:
                program['duration_minutes'] = 30  # Default duration
            
            # Remove extra fields that aren't in the Program model
            extra_fields = ['duration', 'end_time', 'is_recurring', 'new_program', 'updated_at']
            for field in extra_fields:
                if field in program:
                    del program[field]
            
            # Convert datetime strings to datetime objects if needed
            if 'created_at' in program and isinstance(program['created_at'], str):
                from datetime import datetime
                program['created_at'] = datetime.fromisoformat(program['created_at'].replace('Z', '+00:00'))
            
            day = program["day_of_week"]
            if day not in schedule:
                schedule[day] = []
            schedule[day].append(Program(**program))
        except Exception as e:
            print(f"Error converting program {program.get('id', 'unknown')} for schedule: {e}")
            continue
    
    return schedule

# Impact Stories endpoints
@api_router.post("/impact-stories", response_model=ImpactStory)
async def create_impact_story(story: ImpactStoryCreate):
    story_dict = story.dict()
    story_obj = ImpactStory(**story_dict)
    await db.impact_stories.insert_one(story_obj.dict())
    return story_obj

@api_router.get("/impact-stories", response_model=List[ImpactStory])
async def get_impact_stories(featured_only: bool = False):
    filter_dict = {}
    if featured_only:
        filter_dict["is_featured"] = True
    
    stories = await db.impact_stories.find(filter_dict).sort("created_at", -1).to_list(1000)
    return [ImpactStory(**story) for story in stories]

# News endpoints
@api_router.post("/news", response_model=NewsUpdate)
async def create_news_update(news: NewsUpdateCreate):
    news_dict = news.dict()
    news_obj = NewsUpdate(**news_dict)
    await db.news_updates.insert_one(news_obj.dict())
    return news_obj

@api_router.get("/news", response_model=List[NewsUpdate])
async def get_news_updates(published_only: bool = True):
    filter_dict = {}
    if published_only:
        filter_dict["is_published"] = True
    
    news = await db.news_updates.find(filter_dict).sort("created_at", -1).to_list(1000)
    return [NewsUpdate(**news_item) for news_item in news]

# Contact endpoints
@api_router.post("/contact", response_model=ContactMessage)
async def create_contact_message(message: ContactMessageCreate):
    message_dict = message.dict()
    message_obj = ContactMessage(**message_dict)
    await db.contact_messages.insert_one(message_obj.dict())
    return message_obj

@api_router.get("/contact", response_model=List[ContactMessage])
async def get_contact_messages():
    messages = await db.contact_messages.find().sort("created_at", -1).to_list(1000)
    return [ContactMessage(**message) for message in messages]

# Major Gifts endpoints
@api_router.get("/major-gifts-settings")
async def get_major_gifts_settings():
    """Get major gifts settings for calendar, documents, etc."""
    # In real implementation, this would come from database
    settings = MajorGiftsSettings()
    return settings

@api_router.put("/major-gifts-settings")
async def update_major_gifts_settings(settings: MajorGiftsSettings):
    """Update major gifts settings (admin only in real app)"""
    # In real implementation, this would update the database
    return {"message": "Major gifts settings updated successfully", "settings": settings}

@api_router.get("/payment-settings")
async def get_payment_settings():
    """Get payment settings for Direct Liberia, wire transfers, etc."""
    # In real implementation, this would come from database
    settings = PaymentSettings()
    return settings

@api_router.put("/payment-settings") 
async def update_payment_settings(settings: PaymentSettings):
    """Update payment settings (admin only in real app)"""
    # In real implementation, this would update the database
    return {"message": "Payment settings updated successfully", "settings": settings}

@api_router.post("/pledges")
async def create_pledge(pledge: MajorGiftPledgeCreate):
    """Create a new major gift pledge"""
    try:
        # Create pledge document
        pledge_doc = MajorGiftPledge(**pledge.model_dump())
        
        # Insert into database
        result = await db.major_gift_pledges.insert_one(pledge_doc.model_dump())
        
        # Send email notification (in real app, this would send actual email)
        print(f"New major gift pledge received: {pledge.name} - ${pledge.amount}")
        
        return {
            "message": "Pledge created successfully",
            "pledgeId": pledge_doc.id,
            "status": "success"
        }
    except Exception as e:
        print(f"Error creating pledge: {e}")
        raise HTTPException(status_code=500, detail="Failed to create pledge")

@api_router.get("/pledges")
async def get_pledges():
    """Get all major gift pledges (admin only)"""
    try:
        pledges = await db.major_gift_pledges.find().to_list(1000)
        return pledges
    except Exception as e:
        print(f"Error fetching pledges: {e}")
        return []

# Station Settings endpoints (updated)
@api_router.get("/station-settings")
async def get_station_settings():
    """Get station settings for WhatsApp number, email, etc."""
    # In a real implementation, this would come from database
    # For now, return static settings
    settings = StationSettings()
    return settings

@api_router.put("/station-settings")
async def update_station_settings(settings: StationSettings):
    """Update station settings (admin only in real app)"""
    # In a real implementation, this would update the database
    # For now, just return the submitted settings
    return {"message": "Station settings updated successfully", "settings": settings}

# Newsletter signup endpoint
@api_router.post("/newsletter-signup")
async def newsletter_signup(signup: NewsletterSignupCreate):
    # Store the newsletter signup in database
    signup_record = {
        "email": signup.email,
        "admin_email": signup.adminEmail,
        "subscribed_at": datetime.utcnow(),
        "id": str(uuid.uuid4())
    }
    await db.newsletter_signups.insert_one(signup_record)
    
    # In a real implementation, you would send an email to admin@proudlyliberian.com
    # For now, we'll just store the subscription
    
    return {"message": "Successfully subscribed to newsletter", "email": signup.email}

# Church Partners endpoints
@api_router.get("/church-partners")
async def get_church_partners(country: Optional[str] = None, city: Optional[str] = None, published_only: bool = True):
    try:
        # Test database connection
        total_count = await db.church_partners.count_documents({})
        print(f"Total partners in DB: {total_count}")
        
        query = {}
        if country:
            query["country"] = country
        if city:
            query["city"] = city
        if published_only:
            query["isPublished"] = True
        
        print(f"Query: {query}")  # Debug
        
        partners = await db.church_partners.find(query).to_list(1000)
        print(f"Found {len(partners)} partners")  # Debug
        
        # Sort by sortOrder if provided, then by pastorName
        def sort_key(partner):
            sort_order = partner.get("sortOrder")
            if sort_order is None:
                sort_order = 9999
            return (sort_order, partner.get("pastorName", ""))
        
        partners.sort(key=sort_key)
        
        # Convert MongoDB documents to proper format
        result = []
        for partner in partners:
            # Convert MongoDB _id to string and remove it
            if '_id' in partner:
                del partner['_id']
            # Convert datetime to string if present
            if 'created_at' in partner:
                partner['created_at'] = partner['created_at'].isoformat()
            result.append(partner)
        
        return result
    except Exception as e:
        print(f"Error in get_church_partners: {e}")
        return []

@api_router.post("/church-partners", response_model=ChurchPartner)
async def create_church_partner(partner: ChurchPartnerCreate):
    partner_dict = partner.dict()
    partner_obj = ChurchPartner(**partner_dict)
    await db.church_partners.insert_one(partner_obj.dict())
    return partner_obj

@api_router.get("/church-partners/{partner_id}", response_model=ChurchPartner)
async def get_church_partner(partner_id: str):
    partner = await db.church_partners.find_one({"id": partner_id})
    if not partner:
        raise HTTPException(status_code=404, detail="Church partner not found")
    return ChurchPartner(**partner)

@api_router.put("/church-partners/{partner_id}", response_model=ChurchPartner)
async def update_church_partner(partner_id: str, partner: ChurchPartnerCreate):
    partner_dict = partner.dict()
    partner_dict["id"] = partner_id
    
    result = await db.church_partners.replace_one(
        {"id": partner_id}, 
        partner_dict
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Church partner not found")
    
    return ChurchPartner(**partner_dict)

@api_router.delete("/church-partners/{partner_id}")
async def delete_church_partner(partner_id: str):
    result = await db.church_partners.delete_one({"id": partner_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Church partner not found")
    return {"message": "Church partner deleted successfully"}

# About Page Settings endpoints
@api_router.get("/about-page-settings")
async def get_about_page_settings():
    """Get About page settings for vision story, timeline, and documents with preview images"""
    try:
        # Get base settings
        settings = AboutPageSettings()
        
        # Generate document previews
        preview_urls = await generate_document_previews()
        
        # Update settings with preview images
        settings.radioProjectPreviewImages = preview_urls["radioProjectPreviewImages"]
        settings.maruRadioProposalPreviewImages = preview_urls["maruRadioProposalPreviewImages"]
        
        return settings
    except Exception as e:
        print(f"Error fetching about page settings: {e}")
        # Return settings without previews if there's an error
        settings = AboutPageSettings()
        return settings

@api_router.put("/about-page-settings")
async def update_about_page_settings(settings: AboutPageSettings):
    """Update About page settings (admin only in real app)"""
    try:
        # In a real implementation, this would update the database
        # For now, just return the submitted settings
        return {"message": "About page settings updated successfully", "settings": settings}
    except Exception as e:
        print(f"Error updating about page settings: {e}")
        raise HTTPException(status_code=500, detail="Failed to update about page settings")

# Media endpoints
@api_router.get("/media/videos")
async def get_media_videos():
    """Get all media videos ordered by order field"""
    try:
        # Return default videos for now - in real implementation would come from database
        default_videos = [
            {
                "id": "video-1",
                "title": "From Calling to Airwaves: The Vision Behind Kioo Radio 98.1FM",
                "youtubeUrl": "https://youtu.be/qdYg94D1hn0",
                "language": "en",
                "description": "The inspiring story of God's calling to establish Kioo Radio",
                "order": 1,
                "featured": True
            },
            {
                "id": "video-2", 
                "title": "De l'appel aux ondes : la vision à l'origine de Kioo Radio 98.1FM (French Version)",
                "youtubeUrl": "https://youtu.be/4DaV2C2M074",
                "language": "fr",
                "description": "L'histoire inspirante de l'appel de Dieu à établir Kioo Radio",
                "order": 2,
                "featured": True
            },
            {
                "id": "video-3",
                "title": "From Capitol to Betche Hill, Foya: President Joseph N. Boakai, Sr. Supports Kioo Radio", 
                "youtubeUrl": "https://youtu.be/KmH0jOBGEe8",
                "language": "en",
                "description": "Presidential support for Kioo Radio's mission in Liberia",
                "order": 3,
                "featured": True
            },
            {
                "id": "video-4",
                "title": "Foya City Mayor Hon. Josiah Saahkeh endorses Kioo Radio",
                "youtubeUrl": "https://youtu.be/1flOExdkaDU",
                "language": "en", 
                "description": "Local government support from Foya City Mayor for Kioo Radio",
                "order": 4,
                "featured": True
            }
        ]
        return default_videos
    except Exception as e:
        print(f"Error fetching media videos: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch media videos")

@api_router.get("/media/photos")
async def get_media_photos():
    """Get all media photos ordered by order field"""
    try:
        # Updated with real photos from user uploads
        default_photos = [
            {
                "id": "photo-1",
                "title": "Kioo Radio Studio Construction",
                "imageUrl": "/assets/images/gallery/studio-construction.jpg",
                "alt": "Construction of Kioo Radio studio in Foya",
                "caption": "Building the future of Christian broadcasting in the Makona River Region",
                "location": "Foya, Lofa County, Liberia",
                "featured": True,
                "order": 1
            },
            {
                "id": "photo-2",
                "title": "Community Gathering",
                "imageUrl": "/assets/images/gallery/community-gathering.jpg", 
                "alt": "Community members gathered for radio station blessing",
                "caption": "Local church leaders and community members pray for Kioo Radio",
                "location": "Betche Hill, Foya",
                "featured": True,
                "order": 2
            },
            {
                "id": "photo-3",
                "title": "Broadcasting Equipment",
                "imageUrl": "/assets/images/gallery/broadcasting-equipment.jpg",
                "alt": "Professional radio broadcasting equipment",
                "caption": "State-of-the-art equipment ready to serve the tri-border region",
                "location": "Kioo Radio Studio",
                "featured": True,
                "order": 3
            },
            {
                "id": "photo-4",
                "title": "Radio Tower",
                "imageUrl": "/assets/images/gallery/radio-tower.jpg",
                "alt": "Kioo Radio transmission tower",
                "caption": "The tower that will broadcast hope across the Makona River Region",
                "location": "Foya, Lofa County, Liberia",
                "featured": False,
                "order": 4
            }
        ]
        return default_photos
    except Exception as e:
        print(f"Error fetching media photos: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch media photos")

@api_router.get("/media/settings")
async def get_media_settings():
    """Get media settings for display configuration"""
    try:
        settings = MediaSettings()
        return settings
    except Exception as e:
        print(f"Error fetching media settings: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch media settings")

@api_router.put("/media/settings")
async def update_media_settings(settings: MediaSettings):
    """Update media settings (admin only in real app)"""
    try:
        # In real implementation, this would update the database
        return {"message": "Media settings updated successfully", "settings": settings}
    except Exception as e:
        print(f"Error updating media settings: {e}")
        raise HTTPException(status_code=500, detail="Failed to update media settings")

class LiveBroadcastSchedule(BaseModel):
    country: str
    liveDays: List[str]
    preRecordedDays: List[str]
    specialNote: str
    colorCode: str

# Live Broadcast Schedule endpoints
@api_router.get("/live-broadcast-schedule")
async def get_live_broadcast_schedule():
    """Get live broadcast schedule based on Weekly_Live_vs_Pre-Recorded_Rotation"""
    try:
        # Weekly schedule data with Makona Talk Show special case
        weekly_schedule = {
            "monday": {"liberia": "live", "sierra_leone": "pre-recorded", "guinea": "pre-recorded"},
            "tuesday": {"liberia": "live", "sierra_leone": "live", "guinea": "pre-recorded"},
            "wednesday": {"liberia": "live", "sierra_leone": "pre-recorded", "guinea": "live"},
            "thursday": {"liberia": "live", "sierra_leone": "pre-recorded", "guinea": "pre-recorded"},
            "friday": {"liberia": "live", "sierra_leone": "live", "guinea": "pre-recorded"},
            "saturday": {
                "liberia": "live", 
                "sierra_leone": "live", 
                "guinea": "live",
                "special_program": {
                    "name": "Makona Talk Show",
                    "time": "6:00 AM - 9:00 AM",
                    "all_countries_live": True,
                    "note": "All presenters from Liberia, Sierra Leone, and Guinea are LIVE together"
                }
            },
            "sunday": {"liberia": "live", "sierra_leone": "rotation", "guinea": "rotation", "note": "Sunday service rotates weekly between all three countries"}
        }
        
        # Country summary data
        country_schedules = [
            {
                "country": "Liberia",
                "liveDays": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
                "preRecordedDays": [],
                "specialNote": "Always anchor & daily presence",
                "colorCode": "green"
            },
            {
                "country": "Sierra Leone", 
                "liveDays": ["Tuesday", "Friday", "Saturday"],
                "preRecordedDays": ["Monday", "Wednesday", "Thursday"],
                "specialNote": "Sunday live if in rotation, Saturday morning live for Makona Talk Show",
                "colorCode": "blue"
            },
            {
                "country": "Guinea",
                "liveDays": ["Wednesday", "Saturday"], 
                "preRecordedDays": ["Monday", "Tuesday", "Thursday", "Friday"],
                "specialNote": "Sunday live if in rotation, Saturday morning live for Makona Talk Show",
                "colorCode": "gold"
            }
        ]
        
        return {
            "weeklySchedule": weekly_schedule,
            "countrySchedules": country_schedules,
            "introText": "Because of cross-border travel distances, not every team can be live in Foya every day. To ensure fairness and inclusion, each country has specific live days, while pre-recorded programs fill the gaps. Here is our official weekly live broadcast rotation.",
            "specialPrograms": {
                "makona_talk_show": {
                    "name": "Makona Talk Show",
                    "day": "Saturday",
                    "time": "6:00 AM - 9:00 AM",
                    "description": "Special 3-hour program featuring ALL presenters from Liberia, Sierra Leone, and Guinea broadcasting LIVE together",
                    "all_countries_live": True
                }
            }
        }
    except Exception as e:
        print(f"Error fetching live broadcast schedule: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch live broadcast schedule")

class TestimonyLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    name: Optional[str] = None
    location: str
    program: str
    summary: str
    created_at: datetime = Field(default_factory=datetime.utcnow)

class CallLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    time: str
    phone: Optional[str] = None
    summary: str
    category: str  # Testimony, Question, Complaint, Prayer Request
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Dashboard endpoints
@api_router.get("/dashboard/weather")
async def get_dashboard_weather():
    """Get weather data for broadcast areas using Open-Meteo API"""
    try:
        import aiohttp
        
        # City coordinates
        cities = {
            "Foya, Liberia": {"lat": 8.3580, "lon": -10.2049},
            "Koindu, Sierra Leone": {"lat": 8.2804, "lon": -10.6514},
            "Guéckédou, Guinea": {"lat": 8.5674, "lon": -10.1330},
            "Kissidougou, Guinea": {"lat": 9.1885, "lon": -10.0996}
        }
        
        weather_data = {}
        
        async with aiohttp.ClientSession() as session:
            for city_name, coords in cities.items():
                try:
                    # Open-Meteo API call
                    url = f"https://api.open-meteo.com/v1/forecast?latitude={coords['lat']}&longitude={coords['lon']}&current_weather=true&timezone=GMT"
                    
                    async with session.get(url, timeout=10) as response:
                        if response.status == 200:
                            data = await response.json()
                            current = data.get("current_weather", {})
                            
                            # Map weather codes to conditions (WMO Weather interpretation codes)
                            weather_code = current.get("weathercode", 0)
                            condition_map = {
                                0: "Clear sky", 1: "Mainly clear", 2: "Partly cloudy", 3: "Overcast",
                                45: "Fog", 48: "Depositing rime fog", 51: "Light drizzle", 53: "Moderate drizzle",
                                55: "Dense drizzle", 56: "Light freezing drizzle", 57: "Dense freezing drizzle",
                                61: "Slight rain", 63: "Moderate rain", 65: "Heavy rain", 66: "Light freezing rain",
                                67: "Heavy freezing rain", 71: "Slight snow", 73: "Moderate snow", 75: "Heavy snow",
                                77: "Snow grains", 80: "Slight rain showers", 81: "Moderate rain showers",
                                82: "Violent rain showers", 85: "Slight snow showers", 86: "Heavy snow showers",
                                95: "Thunderstorm", 96: "Thunderstorm with hail", 99: "Thunderstorm with heavy hail"
                            }
                            
                            condition = condition_map.get(weather_code, "Unknown")
                            temperature = round(current.get("temperature", 0))
                            
                            weather_data[city_name] = {
                                "temperature": temperature,
                                "condition": condition,
                                "updated": datetime.utcnow().strftime("%Y-%m-%d %H:%M")
                            }
                        else:
                            # Fallback data if API fails
                            weather_data[city_name] = {
                                "temperature": "N/A",
                                "condition": "Weather unavailable",
                                "updated": datetime.utcnow().strftime("%Y-%m-%d %H:%M")
                            }
                except Exception as city_error:
                    print(f"Error fetching weather for {city_name}: {city_error}")
                    # Fallback data for this city
                    weather_data[city_name] = {
                        "temperature": "N/A",
                        "condition": "Weather unavailable",
                        "updated": datetime.utcnow().strftime("%Y-%m-%d %H:%M")
                    }
        
        return weather_data
    except Exception as e:
        print(f"Error fetching weather data: {e}")
        # Return fallback data for all cities
        fallback_data = {}
        cities = ["Foya, Liberia", "Koindu, Sierra Leone", "Guéckédou, Guinea", "Kissidougou, Guinea"]
        for city in cities:
            fallback_data[city] = {
                "temperature": "N/A",
                "condition": "Weather unavailable", 
                "updated": datetime.utcnow().strftime("%Y-%m-%d %H:%M")
            }
        return fallback_data

@api_router.get("/dashboard/schedule")
async def get_dashboard_schedule():
    """Get weekly program schedule"""
    try:
        # Mock schedule data
        schedule_data = [
            {"day": "Monday", "time": "06:00", "program": "Morning Devotion", "presenter": "Rev. Samuel"},
            {"day": "Monday", "time": "08:00", "program": "Kissi News", "presenter": "Marie Camara"},
            {"day": "Monday", "time": "10:00", "program": "Community Talk", "presenter": "Emmanuel Koroma"},
            {"day": "Tuesday", "time": "06:00", "program": "Morning Devotion", "presenter": "Rev. Samuel"},
            {"day": "Tuesday", "time": "08:00", "program": "French Gospel Hour", "presenter": "Pasteur Jean"},
            {"day": "Tuesday", "time": "10:00", "program": "Health Education", "presenter": "Dr. Fatima"},
            {"day": "Wednesday", "time": "06:00", "program": "Morning Devotion", "presenter": "Rev. Samuel"},
            {"day": "Wednesday", "time": "08:00", "program": "Guinea Connection", "presenter": "Amadou Diallo"},
            {"day": "Wednesday", "time": "10:00", "program": "Bible Study", "presenter": "Pastor John"},
            {"day": "Thursday", "time": "06:00", "program": "Morning Devotion", "presenter": "Rev. Samuel"},
            {"day": "Thursday", "time": "08:00", "program": "Community Forum", "presenter": "Sarah Williams"},
            {"day": "Thursday", "time": "10:00", "program": "Music & Praise", "presenter": "Choir Leader"},
            {"day": "Friday", "time": "06:00", "program": "Morning Devotion", "presenter": "Rev. Samuel"},
            {"day": "Friday", "time": "08:00", "program": "Sierra Leone Hour", "presenter": "Moses Kargbo"},
            {"day": "Friday", "time": "10:00", "program": "Youth Forum", "presenter": "David Kolleh"},
            {"day": "Saturday", "time": "08:00", "program": "Family Time", "presenter": "Mrs. Johnson"},
            {"day": "Saturday", "time": "10:00", "program": "Traditional Music", "presenter": "Cultural Team"},
            {"day": "Sunday", "time": "08:00", "program": "Sunday Service", "presenter": "Various Pastors"},
            {"day": "Sunday", "time": "14:00", "program": "Gospel Hour", "presenter": "Rev. Samuel"}
        ]
        return schedule_data
    except Exception as e:
        print(f"Error fetching schedule: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch schedule")

@api_router.get("/dashboard/presenters")
async def get_dashboard_presenters():
    """Get presenters by country"""
    try:
        presenters_data = {
            "liberia": [
                {
                    "name": "Rev. Samuel Johnson",
                    "programs": ["Morning Devotion", "Gospel Hour"],
                    "schedule": "Daily 6:00 AM, Sunday 2:00 PM"
                },
                {
                    "name": "Sarah Williams",
                    "programs": ["Community Forum", "Women's Hour"],
                    "schedule": "Thursday 8:00 AM, Friday 3:00 PM"
                },
                {
                    "name": "David Kolleh",
                    "programs": ["Youth Forum", "Sports Talk"],
                    "schedule": "Friday 10:00 AM, Saturday 4:00 PM"
                }
            ],
            "sierra_leone": [
                {
                    "name": "Emmanuel Koroma",
                    "programs": ["Community Talk", "Technical Hour"],
                    "schedule": "Monday 10:00 AM, Wednesday 3:00 PM"
                },
                {
                    "name": "Moses Kargbo",
                    "programs": ["Sierra Leone Hour", "Cultural Show"],
                    "schedule": "Friday 8:00 AM, Saturday 1:00 PM"
                },
                {
                    "name": "Rev. Philip Tamba Kamara",
                    "programs": ["Sunday Service", "Bible Study"],
                    "schedule": "Sunday 8:00 AM (rotation), Tuesday 7:00 PM"
                }
            ],
            "guinea": [
                {
                    "name": "Marie Camara",
                    "programs": ["Kissi News", "Community Outreach"],
                    "schedule": "Monday 8:00 AM, Thursday 2:00 PM"
                },
                {
                    "name": "Amadou Diallo",
                    "programs": ["Guinea Connection", "French Hour"],
                    "schedule": "Wednesday 8:00 AM, Saturday 10:00 AM"
                },
                {
                    "name": "Pasteur Jean",
                    "programs": ["French Gospel Hour", "Prayer Meeting"],
                    "schedule": "Tuesday 8:00 AM, Thursday 6:00 PM"
                }
            ]
        }
        return presenters_data
    except Exception as e:
        print(f"Error fetching presenters: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch presenters")

# Email notification function
async def send_email_notification(subject: str, content: str, recipient: str = None):
    """Send email notification for dashboard submissions"""
    try:
        # Get email configuration from environment variables
        smtp_server = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.environ.get('SMTP_PORT', '587'))
        sender_email = os.environ.get('SENDER_EMAIL')
        sender_password = os.environ.get('SENDER_PASSWORD')
        default_recipient = os.environ.get('RECIPIENT_EMAIL', 'kiooradiohq@gmail.com')
        
        if not recipient:
            recipient = default_recipient
            
        if not sender_email or not sender_password:
            print("❌ SMTP credentials not configured in environment variables")
            return False
        
        # Create message
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = recipient
        message["Subject"] = subject
        
        # Add body to email
        message.attach(MIMEText(content, "plain"))
        
        # Send email using SMTP
        print(f"📧 Sending email notification to {recipient}...")
        
        # Create SMTP session
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Enable TLS encryption
        server.login(sender_email, sender_password)
        text = message.as_string()
        server.sendmail(sender_email, recipient, text)
        server.quit()
        
        print(f"✅ Email sent successfully to {recipient}")
        print(f"Subject: {subject}")
        
        return True
    except Exception as e:
        print(f"❌ Error sending email notification: {e}")
        # Fallback: log the email content
        print(f"📧 EMAIL NOTIFICATION (FALLBACK LOG) TO {recipient}:")
        print(f"Subject: {subject}")
        print(f"Content: {content}")
        print("="*50)
        return False

@api_router.post("/dashboard/testimony")
async def submit_testimony(testimony: TestimonyLog):
    """Submit a testimony log"""
    try:
        # In production, save to database
        # For now, just return success
        
        # Send email notification
        subject = f"New Testimony Submission - Kioo Radio Dashboard"
        content = f"""
New testimony has been submitted to the Kioo Radio Presenters Dashboard:

Date: {testimony.date}
Name: {testimony.name or 'Anonymous'}
Location: {testimony.location}
Program: {testimony.program}
Summary: {testimony.summary}

Submitted at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---
Kioo Radio Presenters Dashboard
        """
        
        await send_email_notification(subject, content)
        
        return {"message": "Testimony logged successfully", "id": testimony.id}
    except Exception as e:
        print(f"Error submitting testimony: {e}")
        raise HTTPException(status_code=500, detail="Failed to submit testimony")

@api_router.post("/dashboard/call-log")
async def submit_call_log(call: CallLog):
    """Submit a call log"""
    try:
        # In production, save to database
        # For now, just return success
        
        # Send email notification
        subject = f"New Call Log Entry - Kioo Radio Dashboard"
        content = f"""
New phone call has been logged in the Kioo Radio Presenters Dashboard:

Date: {call.date}
Time: {call.time}
Phone Number: {call.phone or 'Not provided'}
Category: {call.category}
Summary: {call.summary}

Submitted at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---
Kioo Radio Presenters Dashboard
        """
        
        await send_email_notification(subject, content)
        
        return {"message": "Call logged successfully", "id": call.id}
    except Exception as e:
        print(f"Error submitting call log: {e}")
        raise HTTPException(status_code=500, detail="Failed to submit call log")

@api_router.get("/dashboard/export")
async def export_dashboard_data():
    """Export all dashboard data as CSV"""
    try:
        # Mock export data - in production, fetch from database
        csv_content = """Type,Date,Time,Location,Program,Name,Phone,Category,Summary
Testimony,2024-09-01,,Foya,Morning Devotion,John Doe,,,"God healed my family"
Call,2024-09-01,08:30,,Community Forum,,+231-77-123456,Prayer Request,"Please pray for my sick mother"
Testimony,2024-09-02,,Koindu,French Hour,Marie,,,"Testimony about answered prayer"
Call,2024-09-02,14:15,,Bible Study,,+232-88-987654,Question,"Question about baptism"
"""
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=kioo-radio-dashboard.csv"}
        )
    except Exception as e:
        print(f"Error exporting data: {e}")
        raise HTTPException(status_code=500, detail="Failed to export data")

# Coverage areas endpoint
@api_router.get("/coverage")
async def get_coverage_areas():
    return {
        "countries": [
            {"name": "Liberia", "coverage": 85, "major_cities": ["Monrovia", "Gbarnga", "Buchanan"]},
            {"name": "Sierra Leone", "coverage": 70, "major_cities": ["Freetown", "Bo", "Kenema"]},
            {"name": "Guinea", "coverage": 60, "major_cities": ["Conakry", "Kankan", "Labé"]}
        ],
        "total_reach": "Over 2 million people",
        "signal_strength": "Strong FM signal covering 150+ mile radius"
    }

# CRM Models
class Contact(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    phone: Optional[str] = None
    organization: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    contact_type: str = "general"  # general, church_partner, donor, volunteer, presenter
    source: str = "manual"  # manual, contact_form, newsletter, church_partner
    notes: Optional[str] = None
    tags: List[str] = []
    last_contact_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ContactCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    organization: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    contact_type: str = "general"
    notes: Optional[str] = None
    tags: List[str] = []

class ContactUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    organization: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    contact_type: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None
    last_contact_date: Optional[datetime] = None

# CSV Import Models
class VisitorRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date_iso: str  # YYYY-MM-DD format
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    country: Optional[str] = None
    county_or_prefecture: Optional[str] = None
    city_town: Optional[str] = None
    program: Optional[str] = None
    language: Optional[str] = None
    testimony: Optional[str] = None
    source: str = "web"  # web/whatsapp/call
    consent_y_n: str = "Y"  # Y/N
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DonationRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date_iso: str  # YYYY-MM-DD format
    donor_name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    country: Optional[str] = None
    method: str  # OrangeMoney/Lonestar/PayPal/Bank
    amount_currency: str  # LRD/USD
    amount: float
    project_code: Optional[str] = None
    note: Optional[str] = None
    receipt_no: Optional[str] = None
    anonymous_y_n: str = "N"  # Y/N
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProjectRecord(BaseModel):
    project_code: str
    name: str
    description_short: Optional[str] = None
    start_date_iso: Optional[str] = None  # YYYY-MM-DD format
    end_date_iso: Optional[str] = None  # YYYY-MM-DD format
    status: str = "planned"  # planned/active/completed
    budget_currency: str = "USD"  # LRD/USD
    budget_amount: Optional[float] = None
    manager: Optional[str] = None
    country: Optional[str] = None
    tags: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FinanceRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date_iso: str  # YYYY-MM-DD format
    type: str  # income/expense
    category: str
    subcategory: Optional[str] = None
    amount_currency: str  # LRD/USD
    amount: float
    method: Optional[str] = None
    reference: Optional[str] = None
    project_code: Optional[str] = None
    notes: Optional[str] = None
    attachment_url: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TaskReminderRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    due_date_iso: str  # YYYY-MM-DD format
    agency: str  # LRA/LTA/MICAT/Business Registry/Other
    description_short: str
    amount_currency: Optional[str] = None  # LRD/USD
    amount: Optional[float] = None
    status: str = "open"  # open/done
    recurrence: str = "one-time"  # annual/one-time
    contact_person: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserRoleRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    role: str  # admin/finance/project/volunteer
    email: EmailStr
    country: Optional[str] = None
    language_default: str = "en"  # en/fr
    phone: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvoiceRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date_iso: str  # YYYY-MM-DD format
    donor_name: str
    contact: Optional[str] = None
    project_code: Optional[str] = None
    amount_currency: str  # LRD/USD
    amount: float
    status: str = "draft"  # draft/sent/paid
    due_date_iso: Optional[str] = None  # YYYY-MM-DD format
    receipt_no: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StoryRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date_iso: str  # YYYY-MM-DD format
    name_or_anonymous: str
    location: Optional[str] = None
    country: Optional[str] = None
    program: Optional[str] = None
    language: Optional[str] = None
    story_text: str
    approved_y_n: str = "N"  # Y/N
    publish_url: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# CSV Import Response Models
class ImportResult(BaseModel):
    success: bool
    imported_count: int
    error_count: int
    errors: List[str] = []
    validation_errors: List[str] = []

class ImportSchedule(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    file_type: str  # visitors, donations, projects, etc.
    cron_expression: str  # e.g., "0 6 * * *" for daily at 6am
    source_url: Optional[str] = None  # for recurring imports from URLs
    last_run: Optional[datetime] = None
    next_run: Optional[datetime] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# CRM endpoints with authentication
security = HTTPBasic()

def authenticate_admin(credentials: HTTPBasicCredentials = Depends(security)):
    """Simple authentication for CRM endpoints"""
    if credentials.username != "admin" or credentials.password != "kioo2025!":
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return credentials.username

@api_router.get("/crm/contacts", response_model=List[Contact])
async def get_contacts(
    contact_type: Optional[str] = None,
    source: Optional[str] = None,
    country: Optional[str] = None,
    limit: int = 100,
    admin: str = Depends(authenticate_admin)
):
    """Get all contacts with optional filtering"""
    try:
        filter_dict = {}
        if contact_type:
            filter_dict["contact_type"] = contact_type
        if source:
            filter_dict["source"] = source
        if country:
            filter_dict["country"] = country
        
        contacts = await db.contacts.find(filter_dict).sort("created_at", -1).limit(limit).to_list(limit)
        
        # Convert MongoDB documents to API format
        result = []
        for contact in contacts:
            if '_id' in contact:
                del contact['_id']
            # Convert datetime strings to datetime objects if needed
            for field in ['created_at', 'updated_at', 'last_contact_date']:
                if field in contact and isinstance(contact[field], str):
                    try:
                        contact[field] = datetime.fromisoformat(contact[field].replace('Z', '+00:00'))
                    except ValueError:
                        pass
            result.append(Contact(**contact))
        
        return result
    except Exception as e:
        logger.error(f"Failed to get contacts: {e}")
        raise HTTPException(status_code=500, detail="Failed to get contacts")

@api_router.post("/crm/contacts", response_model=Contact)
async def create_contact(contact: ContactCreate, admin: str = Depends(authenticate_admin)):
    """Create a new contact"""
    try:
        contact_dict = contact.dict()
        contact_obj = Contact(**contact_dict)
        
        # Check if contact with same email already exists
        existing = await db.contacts.find_one({"email": contact_obj.email})
        if existing:
            raise HTTPException(status_code=400, detail="Contact with this email already exists")
        
        # Prepare for MongoDB storage
        contact_data = contact_obj.dict()
        for field in ['created_at', 'updated_at', 'last_contact_date']:
            if field in contact_data and contact_data[field]:
                contact_data[field] = contact_data[field].isoformat() if isinstance(contact_data[field], datetime) else contact_data[field]
        
        await db.contacts.insert_one(contact_data)
        return contact_obj
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create contact: {e}")
        raise HTTPException(status_code=500, detail="Failed to create contact")

@api_router.get("/crm/contacts/{contact_id}", response_model=Contact)
async def get_contact(contact_id: str, admin: str = Depends(authenticate_admin)):
    """Get a specific contact by ID"""
    try:
        contact = await db.contacts.find_one({"id": contact_id})
        if not contact:
            raise HTTPException(status_code=404, detail="Contact not found")
        
        if '_id' in contact:
            del contact['_id']
        
        # Convert datetime strings to datetime objects if needed
        for field in ['created_at', 'updated_at', 'last_contact_date']:
            if field in contact and isinstance(contact[field], str):
                try:
                    contact[field] = datetime.fromisoformat(contact[field].replace('Z', '+00:00'))
                except ValueError:
                    pass
        
        return Contact(**contact)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get contact: {e}")
        raise HTTPException(status_code=500, detail="Failed to get contact")

@api_router.put("/crm/contacts/{contact_id}", response_model=Contact)
async def update_contact(contact_id: str, contact_update: ContactUpdate, admin: str = Depends(authenticate_admin)):
    """Update an existing contact"""
    try:
        # Get existing contact
        existing_contact = await db.contacts.find_one({"id": contact_id})
        if not existing_contact:
            raise HTTPException(status_code=404, detail="Contact not found")
        
        # Prepare update data
        update_data = contact_update.dict(exclude_unset=True)
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        # Check for email conflicts if email is being updated
        if "email" in update_data and update_data["email"] != existing_contact.get("email"):
            email_conflict = await db.contacts.find_one({"email": update_data["email"], "id": {"$ne": contact_id}})
            if email_conflict:
                raise HTTPException(status_code=400, detail="Contact with this email already exists")
        
        # Update in database
        await db.contacts.update_one(
            {"id": contact_id},
            {"$set": update_data}
        )
        
        # Get updated contact
        updated_contact = await db.contacts.find_one({"id": contact_id})
        if '_id' in updated_contact:
            del updated_contact['_id']
        
        # Convert datetime strings to datetime objects if needed
        for field in ['created_at', 'updated_at', 'last_contact_date']:
            if field in updated_contact and isinstance(updated_contact[field], str):
                try:
                    updated_contact[field] = datetime.fromisoformat(updated_contact[field].replace('Z', '+00:00'))
                except ValueError:
                    pass
        
        return Contact(**updated_contact)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update contact: {e}")
        raise HTTPException(status_code=500, detail="Failed to update contact")

@api_router.delete("/crm/contacts/{contact_id}")
async def delete_contact(contact_id: str, admin: str = Depends(authenticate_admin)):
    """Delete a contact"""
    try:
        result = await db.contacts.delete_one({"id": contact_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Contact not found")
        return {"message": "Contact deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete contact: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete contact")

@api_router.get("/crm/stats")
async def get_crm_stats(admin: str = Depends(authenticate_admin)):
    """Get CRM statistics"""
    try:
        # Contact counts by type
        total_contacts = await db.contacts.count_documents({})
        
        # Count by contact type
        contact_types = await db.contacts.aggregate([
            {"$group": {"_id": "$contact_type", "count": {"$sum": 1}}}
        ]).to_list(100)
        
        # Count by source
        contact_sources = await db.contacts.aggregate([
            {"$group": {"_id": "$source", "count": {"$sum": 1}}}
        ]).to_list(100)
        
        # Count by country
        contact_countries = await db.contacts.aggregate([
            {"$group": {"_id": "$country", "count": {"$sum": 1}}}
        ]).to_list(100)
        
        # Recent contacts (last 30 days)
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        recent_contacts = await db.contacts.count_documents({
            "created_at": {"$gte": thirty_days_ago.isoformat()}
        })
        
        # Newsletter subscribers count
        newsletter_count = await db.newsletter_subscriptions.count_documents({})
        
        # Contact form submissions count
        contact_form_count = await db.contact_form_submissions.count_documents({})
        
        # Church partners count
        church_partners_count = await db.church_partners.count_documents({})
        
        return {
            "total_contacts": total_contacts,
            "recent_contacts": recent_contacts,
            "newsletter_subscribers": newsletter_count,
            "contact_form_submissions": contact_form_count,
            "church_partners": church_partners_count,
            "by_type": {item["_id"]: item["count"] for item in contact_types},
            "by_source": {item["_id"]: item["count"] for item in contact_sources},
            "by_country": {item["_id"]: item["count"] for item in contact_countries}
        }
    except Exception as e:
        logger.error(f"Failed to get CRM stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get CRM stats")

@api_router.post("/crm/import-from-sources")
async def import_contacts_from_sources(admin: str = Depends(authenticate_admin)):
    """Import contacts from existing data sources (newsletter, contact forms, church partners)"""
    try:
        imported_count = 0
        
        # Import from newsletter subscriptions
        newsletter_subs = await db.newsletter_subscriptions.find().to_list(1000)
        for sub in newsletter_subs:
            existing = await db.contacts.find_one({"email": sub.get("email")})
            if not existing and sub.get("email"):
                contact = Contact(
                    name=sub.get("email", "").split("@")[0],  # Use email prefix as name
                    email=sub.get("email"),
                    contact_type="newsletter",
                    source="newsletter",
                    notes=f"Imported from newsletter subscription on {sub.get('timestamp', '')}"
                )
                contact_data = contact.dict()
                for field in ['created_at', 'updated_at', 'last_contact_date']:
                    if field in contact_data and contact_data[field]:
                        contact_data[field] = contact_data[field].isoformat() if isinstance(contact_data[field], datetime) else contact_data[field]
                
                await db.contacts.insert_one(contact_data)
                imported_count += 1
        
        # Import from contact form submissions
        contact_forms = await db.contact_form_submissions.find().to_list(1000)
        for form in contact_forms:
            existing = await db.contacts.find_one({"email": form.get("email")})
            if not existing and form.get("email"):
                contact = Contact(
                    name=form.get("name", form.get("email", "").split("@")[0]),
                    email=form.get("email"),
                    contact_type="general",
                    source="contact_form",
                    notes=f"Contact form: {form.get('subject', '')} - {form.get('message', '')[:100]}..."
                )
                contact_data = contact.dict()
                for field in ['created_at', 'updated_at', 'last_contact_date']:
                    if field in contact_data and contact_data[field]:
                        contact_data[field] = contact_data[field].isoformat() if isinstance(contact_data[field], datetime) else contact_data[field]
                
                await db.contacts.insert_one(contact_data)
                imported_count += 1
        
        # Import from church partners
        church_partners = await db.church_partners.find().to_list(1000)
        for partner in church_partners:
            partner_email = partner.get("pastorEmail") or partner.get("churchEmail")
            if partner_email:
                existing = await db.contacts.find_one({"email": partner_email})
                if not existing:
                    contact = Contact(
                        name=f"Pastor {partner.get('pastorName', 'Unknown')}",
                        email=partner_email,
                        phone=partner.get("contactPhone"),
                        organization=partner.get("churchName"),
                        city=partner.get("city"),
                        country=partner.get("country"),
                        contact_type="church_partner",
                        source="church_partner",
                        notes=f"Church: {partner.get('churchName', '')} - {partner.get('aboutChurch', '')[:100]}..."
                    )
                    contact_data = contact.dict()
                    for field in ['created_at', 'updated_at', 'last_contact_date']:
                        if field in contact_data and contact_data[field]:
                            contact_data[field] = contact_data[field].isoformat() if isinstance(contact_data[field], datetime) else contact_data[field]
                    
                    await db.contacts.insert_one(contact_data)
                    imported_count += 1
        
        return {"message": f"Successfully imported {imported_count} contacts from existing sources"}
    except Exception as e:
        logger.error(f"Failed to import contacts: {e}")
        raise HTTPException(status_code=500, detail="Failed to import contacts")

# CSV Import Helper Functions
def validate_csv_data(df: pd.DataFrame, file_type: str) -> List[str]:
    """Validate CSV data and return list of errors"""
    errors = []
    
    # Define required fields for each file type
    required_fields = {
        'visitors': ['name', 'email'],
        'donations': ['donor_name', 'amount', 'amount_currency'],
        'projects': ['project_code', 'name'],
        'finance': ['type', 'category', 'amount', 'amount_currency'],
        'tasks_reminders': ['due_date_iso', 'agency', 'description_short'],
        'users_roles': ['name', 'role', 'email'],
        'invoices': ['donor_name', 'amount', 'amount_currency'],
        'stories': ['name_or_anonymous', 'story_text']
    }
    
    # Check required fields
    if file_type in required_fields:
        for field in required_fields[file_type]:
            if field not in df.columns:
                errors.append(f"Missing required field: {field}")
    
    # Validate data types and formats
    for index, row in df.iterrows():
        row_num = index + 2  # +2 for header and 0-based indexing
        
        # Validate email format
        if 'email' in row and pd.notna(row['email']):
            if '@' not in str(row['email']):
                errors.append(f"Row {row_num}: Invalid email format: {row['email']}")
        
        # Validate date format (ISO dates should be YYYY-MM-DD)
        date_fields = ['date_iso', 'start_date_iso', 'end_date_iso', 'due_date_iso']
        for date_field in date_fields:
            if date_field in row and pd.notna(row[date_field]):
                try:
                    datetime.strptime(str(row[date_field]), '%Y-%m-%d')
                except ValueError:
                    errors.append(f"Row {row_num}: Invalid date format in {date_field}: {row[date_field]}. Expected YYYY-MM-DD")
        
        # Validate currency codes
        if 'amount_currency' in row and pd.notna(row['amount_currency']):
            if str(row['amount_currency']).upper() not in ['USD', 'LRD']:
                errors.append(f"Row {row_num}: Invalid currency code: {row['amount_currency']}. Must be USD or LRD")
        
        # Validate amounts are numeric
        if 'amount' in row and pd.notna(row['amount']):
            try:
                float(row['amount'])
            except ValueError:
                errors.append(f"Row {row_num}: Amount must be numeric: {row['amount']}")
        
        # Validate Y/N fields
        yn_fields = ['consent_y_n', 'anonymous_y_n', 'approved_y_n']
        for yn_field in yn_fields:
            if yn_field in row and pd.notna(row[yn_field]):
                if str(row[yn_field]).upper() not in ['Y', 'N']:
                    errors.append(f"Row {row_num}: {yn_field} must be Y or N: {row[yn_field]}")
    
    return errors

async def process_csv_import(df: pd.DataFrame, file_type: str) -> ImportResult:
    """Process CSV data and import into appropriate collection"""
    imported_count = 0
    error_count = 0
    errors = []
    
    try:
        collection_map = {
            'visitors': db.visitors,
            'donations': db.donations,
            'projects': db.projects,
            'finance': db.finance,
            'tasks_reminders': db.tasks_reminders,
            'users_roles': db.users_roles,
            'invoices': db.invoices,
            'stories': db.stories
        }
        
        model_map = {
            'visitors': VisitorRecord,
            'donations': DonationRecord,
            'projects': ProjectRecord,
            'finance': FinanceRecord,
            'tasks_reminders': TaskReminderRecord,
            'users_roles': UserRoleRecord,
            'invoices': InvoiceRecord,
            'stories': StoryRecord
        }
        
        if file_type not in collection_map:
            return ImportResult(
                success=False,
                imported_count=0,
                error_count=0,
                errors=[f"Unsupported file type: {file_type}"]
            )
        
        collection = collection_map[file_type]
        model_class = model_map[file_type]
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Convert row to dict and handle NaN values
                row_dict = row.to_dict()
                for key, value in row_dict.items():
                    if pd.isna(value):
                        row_dict[key] = None
                
                # Create model instance
                record = model_class(**row_dict)
                
                # Prepare for MongoDB storage
                record_data = record.dict()
                for field in ['created_at']:
                    if field in record_data and record_data[field]:
                        record_data[field] = record_data[field].isoformat() if isinstance(record_data[field], datetime) else record_data[field]
                
                # Insert into database
                await collection.insert_one(record_data)
                imported_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"Row {index + 2}: {str(e)}")
                logger.error(f"Failed to import row {index + 2}: {e}")
        
        return ImportResult(
            success=error_count == 0,
            imported_count=imported_count,
            error_count=error_count,
            errors=errors
        )
        
    except Exception as e:
        logger.error(f"Failed to process CSV import: {e}")
        return ImportResult(
            success=False,
            imported_count=imported_count,
            error_count=error_count,
            errors=[f"Import processing error: {str(e)}"]
        )

# CSV Import Endpoints
@api_router.post("/crm/import-csv", response_model=ImportResult)
async def import_csv_data(
    file_type: str = Form(...),
    csv_content: str = Form(...),
    admin: str = Depends(authenticate_admin)
):
    """Import data from CSV content"""
    try:
        # Validate file type
        valid_types = ['visitors', 'donations', 'projects', 'finance', 'tasks_reminders', 'users_roles', 'invoices', 'stories']
        if file_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid file type. Must be one of: {', '.join(valid_types)}")
        
        # Parse CSV content
        try:
            df = pd.read_csv(StringIO(csv_content))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid CSV format: {str(e)}")
        
        if df.empty:
            raise HTTPException(status_code=400, detail="CSV file is empty")
        
        # Validate CSV data
        validation_errors = validate_csv_data(df, file_type)
        if validation_errors:
            return ImportResult(
                success=False,
                imported_count=0,
                error_count=len(validation_errors),
                validation_errors=validation_errors
            )
        
        # Process import
        result = await process_csv_import(df, file_type)
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to import CSV data: {e}")
        raise HTTPException(status_code=500, detail="Failed to import CSV data")

@api_router.get("/crm/import-history")
async def get_import_history(admin: str = Depends(authenticate_admin)):
    """Get import history and statistics"""
    try:
        # Get counts from each collection
        collections = {
            'visitors': db.visitors,
            'donations': db.donations,
            'projects': db.projects,
            'finance': db.finance,
            'tasks_reminders': db.tasks_reminders,
            'users_roles': db.users_roles,
            'invoices': db.invoices,
            'stories': db.stories
        }
        
        stats = {}
        for collection_name, collection in collections.items():
            count = await collection.count_documents({})
            recent_count = await collection.count_documents({
                "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()}
            })
            stats[collection_name] = {
                "total_records": count,
                "recent_records": recent_count
            }
        
        return {
            "import_statistics": stats,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        
    except Exception as e:
        logger.error(f"Failed to get import history: {e}")
        raise HTTPException(status_code=500, detail="Failed to get import history")

@api_router.post("/crm/schedules", response_model=ImportSchedule)
async def create_import_schedule(schedule: ImportSchedule, admin: str = Depends(authenticate_admin)):
    """Create a scheduled import job"""
    try:
        # Validate cron expression
        try:
            CronTrigger.from_crontab(schedule.cron_expression)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Invalid cron expression: {str(e)}")
        
        # Save schedule to database
        schedule_data = schedule.dict()
        for field in ['created_at', 'last_run', 'next_run']:
            if field in schedule_data and schedule_data[field]:
                schedule_data[field] = schedule_data[field].isoformat() if isinstance(schedule_data[field], datetime) else schedule_data[field]
        
        await db.import_schedules.insert_one(schedule_data)
        return schedule
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create import schedule: {e}")
        raise HTTPException(status_code=500, detail="Failed to create import schedule")

@api_router.get("/crm/schedules", response_model=List[ImportSchedule])
async def get_import_schedules(admin: str = Depends(authenticate_admin)):
    """Get all import schedules"""
    try:
        schedules = await db.import_schedules.find().to_list(100)
        
        result = []
        for schedule in schedules:
            if '_id' in schedule:
                del schedule['_id']
            # Convert datetime strings to datetime objects if needed
            for field in ['created_at', 'last_run', 'next_run']:
                if field in schedule and isinstance(schedule[field], str):
                    try:
                        schedule[field] = datetime.fromisoformat(schedule[field].replace('Z', '+00:00'))
                    except ValueError:
                        pass
            result.append(ImportSchedule(**schedule))
        
        return result
        
    except Exception as e:
        logger.error(f"Failed to get import schedules: {e}")
        raise HTTPException(status_code=500, detail="Failed to get import schedules")

@api_router.delete("/crm/schedules/{schedule_id}")
async def delete_import_schedule(schedule_id: str, admin: str = Depends(authenticate_admin)):
    """Delete an import schedule"""
    try:
        result = await db.import_schedules.delete_one({"id": schedule_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Schedule not found")
        return {"message": "Schedule deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete import schedule: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete import schedule")

# Visitors Management Models
class VisitorCreate(BaseModel):
    date_iso: str
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    country: str
    county_or_prefecture: Optional[str] = None
    city_town: Optional[str] = None
    program: str
    language: Optional[str] = None
    testimony: str
    source: str = "web"
    consent_y_n: str = "Y"

class VisitorUpdate(BaseModel):
    date_iso: Optional[str] = None
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    country: Optional[str] = None
    county_or_prefecture: Optional[str] = None
    city_town: Optional[str] = None
    program: Optional[str] = None
    language: Optional[str] = None
    testimony: Optional[str] = None
    source: Optional[str] = None
    consent_y_n: Optional[str] = None

# Visitors Management Endpoints
@api_router.get("/visitors", response_model=List[VisitorRecord])
async def get_visitors(
    month: Optional[str] = None,  # YYYY-MM format
    country: Optional[str] = None,
    program: Optional[str] = None,
    source: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    admin: str = Depends(authenticate_admin)
):
    """Get visitors with optional filtering"""
    try:
        filter_dict = {}
        
        # Month filter (YYYY-MM format)
        if month:
            try:
                year, month_num = month.split('-')
                start_date = f"{year}-{month_num.zfill(2)}-01"
                # Calculate end date (first day of next month)
                if month_num == "12":
                    end_date = f"{int(year) + 1}-01-01"
                else:
                    end_date = f"{year}-{str(int(month_num) + 1).zfill(2)}-01"
                
                filter_dict["date_iso"] = {
                    "$gte": start_date,
                    "$lt": end_date
                }
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        # Other filters
        if country:
            filter_dict["country"] = country
        if program:
            filter_dict["program"] = program
        if source:
            filter_dict["source"] = source
        
        # Get visitors with pagination
        visitors = await db.visitors.find(filter_dict).sort("date_iso", -1).skip(skip).limit(limit).to_list(limit)
        
        # Convert MongoDB documents to API format
        result = []
        for visitor in visitors:
            if '_id' in visitor:
                del visitor['_id']
            # Convert datetime strings to datetime objects if needed
            if 'created_at' in visitor and isinstance(visitor['created_at'], str):
                try:
                    visitor['created_at'] = datetime.fromisoformat(visitor['created_at'].replace('Z', '+00:00'))
                except ValueError:
                    pass
            result.append(VisitorRecord(**visitor))
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get visitors: {e}")
        raise HTTPException(status_code=500, detail="Failed to get visitors")

@api_router.post("/visitors", response_model=VisitorRecord)
async def create_visitor(visitor: VisitorCreate, admin: str = Depends(authenticate_admin)):
    """Create a new visitor"""
    try:
        # Validate date format
        try:
            datetime.strptime(visitor.date_iso, '%Y-%m-%d')
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Validate consent
        if visitor.consent_y_n not in ['Y', 'N']:
            raise HTTPException(status_code=400, detail="Consent must be Y or N")
        
        # Create visitor record
        visitor_dict = visitor.dict()
        visitor_obj = VisitorRecord(**visitor_dict)
        
        # Prepare for MongoDB storage
        visitor_data = visitor_obj.dict()
        for field in ['created_at']:
            if field in visitor_data and visitor_data[field]:
                visitor_data[field] = visitor_data[field].isoformat() if isinstance(visitor_data[field], datetime) else visitor_data[field]
        
        await db.visitors.insert_one(visitor_data)
        return visitor_obj
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create visitor: {e}")
        raise HTTPException(status_code=500, detail="Failed to create visitor")

@api_router.get("/visitors/{visitor_id}", response_model=VisitorRecord)
async def get_visitor(visitor_id: str, admin: str = Depends(authenticate_admin)):
    """Get a specific visitor by ID"""
    try:
        visitor = await db.visitors.find_one({"id": visitor_id})
        if not visitor:
            raise HTTPException(status_code=404, detail="Visitor not found")
        
        if '_id' in visitor:
            del visitor['_id']
        
        # Convert datetime strings to datetime objects if needed
        if 'created_at' in visitor and isinstance(visitor['created_at'], str):
            try:
                visitor['created_at'] = datetime.fromisoformat(visitor['created_at'].replace('Z', '+00:00'))
            except ValueError:
                pass
        
        return VisitorRecord(**visitor)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get visitor: {e}")
        raise HTTPException(status_code=500, detail="Failed to get visitor")

@api_router.put("/visitors/{visitor_id}", response_model=VisitorRecord)
async def update_visitor(visitor_id: str, visitor_update: VisitorUpdate, admin: str = Depends(authenticate_admin)):
    """Update an existing visitor"""
    try:
        # Get existing visitor
        existing_visitor = await db.visitors.find_one({"id": visitor_id})
        if not existing_visitor:
            raise HTTPException(status_code=404, detail="Visitor not found")
        
        # Prepare update data
        update_data = visitor_update.dict(exclude_unset=True)
        
        # Validate date format if provided
        if "date_iso" in update_data and update_data["date_iso"]:
            try:
                datetime.strptime(update_data["date_iso"], '%Y-%m-%d')
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Validate consent if provided
        if "consent_y_n" in update_data and update_data["consent_y_n"] not in ['Y', 'N']:
            raise HTTPException(status_code=400, detail="Consent must be Y or N")
        
        # Update in database
        await db.visitors.update_one(
            {"id": visitor_id},
            {"$set": update_data}
        )
        
        # Get updated visitor
        updated_visitor = await db.visitors.find_one({"id": visitor_id})
        if '_id' in updated_visitor:
            del updated_visitor['_id']
        
        # Convert datetime strings to datetime objects if needed
        if 'created_at' in updated_visitor and isinstance(updated_visitor['created_at'], str):
            try:
                updated_visitor['created_at'] = datetime.fromisoformat(updated_visitor['created_at'].replace('Z', '+00:00'))
            except ValueError:
                pass
        
        return VisitorRecord(**updated_visitor)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update visitor: {e}")
        raise HTTPException(status_code=500, detail="Failed to update visitor")

@api_router.delete("/visitors/{visitor_id}")
async def delete_visitor(visitor_id: str, admin: str = Depends(authenticate_admin)):
    """Delete a visitor"""
    try:
        result = await db.visitors.delete_one({"id": visitor_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Visitor not found")
        return {"message": "Visitor deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete visitor: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete visitor")

@api_router.get("/visitors/export/csv")
async def export_visitors_csv(
    month: Optional[str] = None,
    country: Optional[str] = None,
    program: Optional[str] = None,
    source: Optional[str] = None,
    admin: str = Depends(authenticate_admin)
):
    """Export visitors as CSV"""
    try:
        # Build filter
        filter_dict = {}
        if month:
            try:
                year, month_num = month.split('-')
                start_date = f"{year}-{month_num.zfill(2)}-01"
                if month_num == "12":
                    end_date = f"{int(year) + 1}-01-01"
                else:
                    end_date = f"{year}-{str(int(month_num) + 1).zfill(2)}-01"
                filter_dict["date_iso"] = {"$gte": start_date, "$lt": end_date}
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        if country:
            filter_dict["country"] = country
        if program:
            filter_dict["program"] = program
        if source:
            filter_dict["source"] = source
        
        # Get all matching visitors
        visitors = await db.visitors.find(filter_dict).sort("date_iso", -1).to_list(10000)
        
        # Create CSV content
        csv_content = "date_iso,name,phone,email,country,county_or_prefecture,city_town,program,language,testimony,source,consent_y_n\n"
        
        for visitor in visitors:
            row = [
                visitor.get("date_iso", ""),
                visitor.get("name", ""),
                visitor.get("phone", ""),
                visitor.get("email", ""),
                visitor.get("country", ""),
                visitor.get("county_or_prefecture", ""),
                visitor.get("city_town", ""),
                visitor.get("program", ""),
                visitor.get("language", ""),
                visitor.get("testimony", "").replace('"', '""'),  # Escape quotes
                visitor.get("source", ""),
                visitor.get("consent_y_n", "")
            ]
            csv_content += '"' + '","'.join(row) + '"\n'
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"visitors_export_{timestamp}.csv"
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export visitors CSV: {e}")
        raise HTTPException(status_code=500, detail="Failed to export visitors CSV")

@api_router.get("/visitors/export/xlsx")
async def export_visitors_xlsx(
    month: Optional[str] = None,
    country: Optional[str] = None,
    program: Optional[str] = None,
    source: Optional[str] = None,
    admin: str = Depends(authenticate_admin)
):
    """Export visitors as XLSX"""
    try:
        # Build filter
        filter_dict = {}
        if month:
            try:
                year, month_num = month.split('-')
                start_date = f"{year}-{month_num.zfill(2)}-01"
                if month_num == "12":
                    end_date = f"{int(year) + 1}-01-01"
                else:
                    end_date = f"{year}-{str(int(month_num) + 1).zfill(2)}-01"
                filter_dict["date_iso"] = {"$gte": start_date, "$lt": end_date}
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        if country:
            filter_dict["country"] = country
        if program:
            filter_dict["program"] = program
        if source:
            filter_dict["source"] = source
        
        # Get all matching visitors
        visitors = await db.visitors.find(filter_dict).sort("date_iso", -1).to_list(10000)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitors"
        
        # Headers
        headers = ["Date", "Name", "Phone", "Email", "Country", "County/Prefecture", "City/Town", "Program", "Language", "Testimony", "Source", "Consent"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for visitor in visitors:
            row = [
                visitor.get("date_iso", ""),
                visitor.get("name", ""),
                visitor.get("phone", ""),
                visitor.get("email", ""),
                visitor.get("country", ""),
                visitor.get("county_or_prefecture", ""),
                visitor.get("city_town", ""),
                visitor.get("program", ""),
                visitor.get("language", ""),
                visitor.get("testimony", ""),
                visitor.get("source", ""),
                visitor.get("consent_y_n", "")
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"visitors_export_{timestamp}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export visitors XLSX: {e}")
        raise HTTPException(status_code=500, detail="Failed to export visitors XLSX")

@api_router.get("/visitors/filter-stats")
async def get_visitors_stats(admin: str = Depends(authenticate_admin)):
    """Get visitors statistics for filters"""
    try:
        # Get unique countries
        countries = await db.visitors.distinct("country")
        countries = [c for c in countries if c]  # Remove empty/null values
        
        # Get unique programs
        programs = await db.visitors.distinct("program")
        programs = [p for p in programs if p]  # Remove empty/null values
        
        # Get unique sources
        sources = await db.visitors.distinct("source")
        sources = [s for s in sources if s]  # Remove empty/null values
        
        # Get date range for month filter
        date_pipeline = [
            {"$group": {
                "_id": None,
                "min_date": {"$min": "$date_iso"},
                "max_date": {"$max": "$date_iso"}
            }}
        ]
        date_result = await db.visitors.aggregate(date_pipeline).to_list(1)
        date_range = date_result[0] if date_result else {"min_date": None, "max_date": None}
        
        return {
            "countries": sorted(countries),
            "programs": sorted(programs),
            "sources": sorted(sources),
            "date_range": date_range,
            "total_visitors": await db.visitors.count_documents({})
        }
        
    except Exception as e:
        logger.error(f"Failed to get visitors stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get visitors stats")

# Donations Management Models
class DonationCreate(BaseModel):
    date_iso: str
    donor_name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    country: Optional[str] = None
    method: str  # OrangeMoney/Lonestar/PayPal/Bank
    amount_currency: str  # LRD/USD
    amount: float
    project_code: Optional[str] = None
    note: Optional[str] = None
    receipt_no: Optional[str] = None
    anonymous_y_n: str = "N"

class DonationUpdate(BaseModel):
    date_iso: Optional[str] = None
    donor_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    country: Optional[str] = None
    method: Optional[str] = None
    amount_currency: Optional[str] = None
    amount: Optional[float] = None
    project_code: Optional[str] = None
    note: Optional[str] = None
    receipt_no: Optional[str] = None
    anonymous_y_n: Optional[str] = None

# Donations Management Endpoints
@api_router.get("/donations", response_model=List[DonationRecord])
async def get_donations(
    month: Optional[str] = None,  # YYYY-MM format
    project_code: Optional[str] = None,
    method: Optional[str] = None,
    anonymous: Optional[str] = None,  # Y/N
    limit: int = 100,
    skip: int = 0,
    admin: str = Depends(authenticate_admin)
):
    """Get donations with optional filtering"""
    try:
        filter_dict = {}
        
        # Month filter (YYYY-MM format)
        if month:
            try:
                year, month_num = month.split('-')
                start_date = f"{year}-{month_num.zfill(2)}-01"
                # Calculate end date (first day of next month)
                if month_num == "12":
                    end_date = f"{int(year) + 1}-01-01"
                else:
                    end_date = f"{year}-{str(int(month_num) + 1).zfill(2)}-01"
                
                filter_dict["date_iso"] = {
                    "$gte": start_date,
                    "$lt": end_date
                }
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        # Other filters
        if project_code:
            filter_dict["project_code"] = project_code
        if method:
            filter_dict["method"] = method
        if anonymous:
            filter_dict["anonymous_y_n"] = anonymous
        
        # Get donations with pagination
        donations = await db.donations.find(filter_dict).sort("date_iso", -1).skip(skip).limit(limit).to_list(limit)
        
        # Convert MongoDB documents to API format
        result = []
        for donation in donations:
            if '_id' in donation:
                del donation['_id']
            # Convert datetime strings to datetime objects if needed
            if 'created_at' in donation and isinstance(donation['created_at'], str):
                try:
                    donation['created_at'] = datetime.fromisoformat(donation['created_at'].replace('Z', '+00:00'))
                except ValueError:
                    pass
            result.append(DonationRecord(**donation))
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get donations: {e}")
        raise HTTPException(status_code=500, detail="Failed to get donations")

@api_router.post("/donations", response_model=DonationRecord)
async def create_donation(donation: DonationCreate, admin: str = Depends(authenticate_admin)):
    """Create a new donation"""
    try:
        # Validate date format
        try:
            datetime.strptime(donation.date_iso, '%Y-%m-%d')
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Validate amount
        if donation.amount <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than 0")
            
        # Validate currency
        if donation.amount_currency not in ['USD', 'LRD']:
            raise HTTPException(status_code=400, detail="Currency must be USD or LRD")
        
        # Validate method
        valid_methods = ['OrangeMoney', 'Lonestar', 'PayPal', 'Bank']
        if donation.method not in valid_methods:
            raise HTTPException(status_code=400, detail=f"Method must be one of: {', '.join(valid_methods)}")
        
        # Validate anonymous
        if donation.anonymous_y_n not in ['Y', 'N']:
            raise HTTPException(status_code=400, detail="Anonymous must be Y or N")
        
        # Create donation record
        donation_dict = donation.dict()
        donation_obj = DonationRecord(**donation_dict)
        
        # Prepare for MongoDB storage
        donation_data = donation_obj.dict()
        for field in ['created_at']:
            if field in donation_data and donation_data[field]:
                donation_data[field] = donation_data[field].isoformat() if isinstance(donation_data[field], datetime) else donation_data[field]
        
        await db.donations.insert_one(donation_data)
        return donation_obj
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create donation: {e}")
        raise HTTPException(status_code=500, detail="Failed to create donation")

@api_router.get("/donations/{donation_id}", response_model=DonationRecord)
async def get_donation(donation_id: str, admin: str = Depends(authenticate_admin)):
    """Get a specific donation by ID"""
    try:
        donation = await db.donations.find_one({"id": donation_id})
        if not donation:
            raise HTTPException(status_code=404, detail="Donation not found")
        
        if '_id' in donation:
            del donation['_id']
        
        # Convert datetime strings to datetime objects if needed
        if 'created_at' in donation and isinstance(donation['created_at'], str):
            try:
                donation['created_at'] = datetime.fromisoformat(donation['created_at'].replace('Z', '+00:00'))
            except ValueError:
                pass
        
        return DonationRecord(**donation)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get donation: {e}")
        raise HTTPException(status_code=500, detail="Failed to get donation")

@api_router.put("/donations/{donation_id}", response_model=DonationRecord)
async def update_donation(donation_id: str, donation_update: DonationUpdate, admin: str = Depends(authenticate_admin)):
    """Update an existing donation"""
    try:
        # Get existing donation
        existing_donation = await db.donations.find_one({"id": donation_id})
        if not existing_donation:
            raise HTTPException(status_code=404, detail="Donation not found")
        
        # Prepare update data
        update_data = donation_update.dict(exclude_unset=True)
        
        # Validate date format if provided
        if "date_iso" in update_data and update_data["date_iso"]:
            try:
                datetime.strptime(update_data["date_iso"], '%Y-%m-%d')
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Validate amount if provided
        if "amount" in update_data and update_data["amount"] is not None:
            if update_data["amount"] <= 0:
                raise HTTPException(status_code=400, detail="Amount must be greater than 0")
        
        # Validate currency if provided
        if "amount_currency" in update_data and update_data["amount_currency"]:
            if update_data["amount_currency"] not in ['USD', 'LRD']:
                raise HTTPException(status_code=400, detail="Currency must be USD or LRD")
        
        # Validate method if provided
        if "method" in update_data and update_data["method"]:
            valid_methods = ['OrangeMoney', 'Lonestar', 'PayPal', 'Bank']
            if update_data["method"] not in valid_methods:
                raise HTTPException(status_code=400, detail=f"Method must be one of: {', '.join(valid_methods)}")
        
        # Validate anonymous if provided
        if "anonymous_y_n" in update_data and update_data["anonymous_y_n"] not in ['Y', 'N']:
            raise HTTPException(status_code=400, detail="Anonymous must be Y or N")
        
        # Update in database
        await db.donations.update_one(
            {"id": donation_id},
            {"$set": update_data}
        )
        
        # Get updated donation
        updated_donation = await db.donations.find_one({"id": donation_id})
        if '_id' in updated_donation:
            del updated_donation['_id']
        
        # Convert datetime strings to datetime objects if needed
        if 'created_at' in updated_donation and isinstance(updated_donation['created_at'], str):
            try:
                updated_donation['created_at'] = datetime.fromisoformat(updated_donation['created_at'].replace('Z', '+00:00'))
            except ValueError:
                pass
        
        return DonationRecord(**updated_donation)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update donation: {e}")
        raise HTTPException(status_code=500, detail="Failed to update donation")

@api_router.delete("/donations/{donation_id}")
async def delete_donation(donation_id: str, admin: str = Depends(authenticate_admin)):
    """Delete a donation"""
    try:
        result = await db.donations.delete_one({"id": donation_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Donation not found")
        return {"message": "Donation deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete donation: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete donation")

@api_router.get("/donations/totals/summary")
async def get_donation_totals(admin: str = Depends(authenticate_admin)):
    """Get donation totals for this month and YTD"""
    try:
        now = datetime.now(timezone.utc)
        
        # This month totals
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_start_str = month_start.strftime('%Y-%m-%d')
        
        # Next month start for range
        if now.month == 12:
            next_month_start = now.replace(year=now.year + 1, month=1, day=1)
        else:
            next_month_start = now.replace(month=now.month + 1, day=1)
        next_month_start_str = next_month_start.strftime('%Y-%m-%d')
        
        # Year start
        year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        year_start_str = year_start.strftime('%Y-%m-%d')
        
        # This month aggregation
        month_pipeline = [
            {"$match": {
                "date_iso": {
                    "$gte": month_start_str,
                    "$lt": next_month_start_str
                }
            }},
            {"$group": {
                "_id": "$amount_currency",
                "total": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }}
        ]
        
        month_results = await db.donations.aggregate(month_pipeline).to_list(10)
        
        # YTD aggregation
        ytd_pipeline = [
            {"$match": {
                "date_iso": {"$gte": year_start_str}
            }},
            {"$group": {
                "_id": "$amount_currency",
                "total": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }}
        ]
        
        ytd_results = await db.donations.aggregate(ytd_pipeline).to_list(10)
        
        # Format results
        month_totals = {"USD": 0, "LRD": 0, "count": 0}
        for result in month_results:
            currency = result["_id"] or "USD"
            month_totals[currency] = result["total"]
            month_totals["count"] += result["count"]
        
        ytd_totals = {"USD": 0, "LRD": 0, "count": 0}
        for result in ytd_results:
            currency = result["_id"] or "USD"
            ytd_totals[currency] = result["total"]
            ytd_totals["count"] += result["count"]
        
        return {
            "month": {
                "period": f"{now.strftime('%B %Y')}",
                "usd_total": month_totals["USD"],
                "lrd_total": month_totals["LRD"],
                "total_donations": month_totals["count"]
            },
            "ytd": {
                "period": f"{now.year}",
                "usd_total": ytd_totals["USD"],
                "lrd_total": ytd_totals["LRD"],
                "total_donations": ytd_totals["count"]
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to get donation totals: {e}")
        raise HTTPException(status_code=500, detail="Failed to get donation totals")

@api_router.get("/donations/export/csv")
async def export_donations_csv(
    month: Optional[str] = None,
    project_code: Optional[str] = None,
    method: Optional[str] = None,
    anonymous: Optional[str] = None,
    admin: str = Depends(authenticate_admin)
):
    """Export donations as CSV"""
    try:
        # Build filter
        filter_dict = {}
        if month:
            try:
                year, month_num = month.split('-')
                start_date = f"{year}-{month_num.zfill(2)}-01"
                if month_num == "12":
                    end_date = f"{int(year) + 1}-01-01"
                else:
                    end_date = f"{year}-{str(int(month_num) + 1).zfill(2)}-01"
                filter_dict["date_iso"] = {"$gte": start_date, "$lt": end_date}
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        if project_code:
            filter_dict["project_code"] = project_code
        if method:
            filter_dict["method"] = method
        if anonymous:
            filter_dict["anonymous_y_n"] = anonymous
        
        # Get all matching donations
        donations = await db.donations.find(filter_dict).sort("date_iso", -1).to_list(10000)
        
        # Create CSV content
        csv_content = "date_iso,donor_name,phone,email,country,method,amount_currency,amount,project_code,note,receipt_no,anonymous_y_n\n"
        
        for donation in donations:
            row = [
                donation.get("date_iso", ""),
                donation.get("donor_name", ""),
                donation.get("phone", ""),
                donation.get("email", ""),
                donation.get("country", ""),
                donation.get("method", ""),
                donation.get("amount_currency", ""),
                str(donation.get("amount", "")),
                donation.get("project_code", ""),
                donation.get("note", "").replace('"', '""'),  # Escape quotes
                donation.get("receipt_no", ""),
                donation.get("anonymous_y_n", "")
            ]
            csv_content += '"' + '","'.join(row) + '"\n'
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"donations_export_{timestamp}.csv"
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export donations CSV: {e}")
        raise HTTPException(status_code=500, detail="Failed to export donations CSV")

@api_router.get("/donations/export/xlsx")
async def export_donations_xlsx(
    month: Optional[str] = None,
    project_code: Optional[str] = None,
    method: Optional[str] = None,
    anonymous: Optional[str] = None,
    admin: str = Depends(authenticate_admin)
):
    """Export donations as XLSX"""
    try:
        # Build filter
        filter_dict = {}
        if month:
            try:
                year, month_num = month.split('-')
                start_date = f"{year}-{month_num.zfill(2)}-01"
                if month_num == "12":
                    end_date = f"{int(year) + 1}-01-01"
                else:
                    end_date = f"{year}-{str(int(month_num) + 1).zfill(2)}-01"
                filter_dict["date_iso"] = {"$gte": start_date, "$lt": end_date}
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        if project_code:
            filter_dict["project_code"] = project_code
        if method:
            filter_dict["method"] = method
        if anonymous:
            filter_dict["anonymous_y_n"] = anonymous
        
        # Get all matching donations
        donations = await db.donations.find(filter_dict).sort("date_iso", -1).to_list(10000)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Donations"
        
        # Headers
        headers = ["Date", "Donor Name", "Phone", "Email", "Country", "Method", "Currency", "Amount", "Project Code", "Note", "Receipt No", "Anonymous"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for donation in donations:
            row = [
                donation.get("date_iso", ""),
                donation.get("donor_name", ""),
                donation.get("phone", ""),
                donation.get("email", ""),
                donation.get("country", ""),
                donation.get("method", ""),
                donation.get("amount_currency", ""),
                donation.get("amount", ""),
                donation.get("project_code", ""),
                donation.get("note", ""),
                donation.get("receipt_no", ""),
                donation.get("anonymous_y_n", "")
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"donations_export_{timestamp}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export donations XLSX: {e}")
        raise HTTPException(status_code=500, detail="Failed to export donations XLSX")

@api_router.get("/donations/management-stats")
async def get_donations_filter_stats(admin: str = Depends(authenticate_admin)):
    """Get donation statistics for filters"""
    try:
        # Get unique project codes
        project_codes = await db.donations.distinct("project_code")
        project_codes = [p for p in project_codes if p]  # Remove empty/null values
        
        # Get unique methods
        methods = await db.donations.distinct("method")
        methods = [m for m in methods if m]  # Remove empty/null values
        
        # Get date range for month filter
        date_pipeline = [
            {"$group": {
                "_id": None,
                "min_date": {"$min": "$date_iso"},
                "max_date": {"$max": "$date_iso"}
            }}
        ]
        date_result = await db.donations.aggregate(date_pipeline).to_list(1)
        date_range = date_result[0] if date_result else {"min_date": None, "max_date": None}
        
        return {
            "project_codes": sorted(project_codes),
            "methods": sorted(methods),
            "date_range": date_range,
            "total_donations": await db.donations.count_documents({})
        }
        
    except Exception as e:
        logger.error(f"Failed to get donations filter stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get donations filter stats")

# Projects Management Models  
class ProjectCreate(BaseModel):
    project_code: str
    name: str
    description_short: Optional[str] = None
    start_date_iso: Optional[str] = None
    end_date_iso: Optional[str] = None
    status: str = "planned"  # planned/active/completed/on-hold/cancelled
    budget_currency: str = "USD"
    budget_amount: Optional[float] = None
    manager: Optional[str] = None
    country: Optional[str] = None
    tags: Optional[str] = None

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description_short: Optional[str] = None
    start_date_iso: Optional[str] = None
    end_date_iso: Optional[str] = None
    status: Optional[str] = None
    budget_currency: Optional[str] = None
    budget_amount: Optional[float] = None
    manager: Optional[str] = None
    country: Optional[str] = None
    tags: Optional[str] = None

# Projects Management Endpoints
@api_router.get("/projects", response_model=List[ProjectRecord])
async def get_projects(
    status: Optional[str] = None,
    country: Optional[str] = None,
    manager: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    admin: str = Depends(authenticate_admin)
):
    """Get projects with optional filtering"""
    try:
        filter_dict = {}
        
        if status:
            filter_dict["status"] = status
        if country:
            filter_dict["country"] = country
        if manager:
            filter_dict["manager"] = manager
        
        # Get projects with pagination
        projects = await db.projects.find(filter_dict).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
        
        # Convert MongoDB documents to API format
        result = []
        for project in projects:
            if '_id' in project:
                del project['_id']
            # Convert datetime strings to datetime objects if needed
            if 'created_at' in project and isinstance(project['created_at'], str):
                try:
                    project['created_at'] = datetime.fromisoformat(project['created_at'].replace('Z', '+00:00'))
                except ValueError:
                    pass
            result.append(ProjectRecord(**project))
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get projects: {e}")
        raise HTTPException(status_code=500, detail="Failed to get projects")

@api_router.post("/projects", response_model=ProjectRecord)
async def create_project(project: ProjectCreate, admin: str = Depends(authenticate_admin)):
    """Create a new project"""
    try:
        # Validate project code uniqueness
        existing_project = await db.projects.find_one({"project_code": project.project_code})
        if existing_project:
            raise HTTPException(status_code=400, detail="Project code already exists")
        
        # Validate dates if provided
        if project.start_date_iso:
            try:
                datetime.strptime(project.start_date_iso, '%Y-%m-%d')
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid start date format. Use YYYY-MM-DD")
        
        if project.end_date_iso:
            try:
                datetime.strptime(project.end_date_iso, '%Y-%m-%d')
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid end date format. Use YYYY-MM-DD")
        
        # Validate status
        valid_statuses = ['planned', 'active', 'completed', 'on-hold', 'cancelled']
        if project.status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Status must be one of: {', '.join(valid_statuses)}")
        
        # Validate currency
        if project.budget_currency not in ['USD', 'LRD']:
            raise HTTPException(status_code=400, detail="Budget currency must be USD or LRD")
        
        # Validate budget amount if provided
        if project.budget_amount is not None and project.budget_amount < 0:
            raise HTTPException(status_code=400, detail="Budget amount must be positive")
        
        # Create project record
        project_dict = project.dict()
        project_obj = ProjectRecord(**project_dict)
        
        # Prepare for MongoDB storage
        project_data = project_obj.dict()
        for field in ['created_at']:
            if field in project_data and project_data[field]:
                project_data[field] = project_data[field].isoformat() if isinstance(project_data[field], datetime) else project_data[field]
        
        await db.projects.insert_one(project_data)
        return project_obj
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create project: {e}")
        raise HTTPException(status_code=500, detail="Failed to create project")

@api_router.get("/projects/{project_code}", response_model=ProjectRecord)
async def get_project(project_code: str, admin: str = Depends(authenticate_admin)):
    """Get a specific project by code"""
    try:
        project = await db.projects.find_one({"project_code": project_code})
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        if '_id' in project:
            del project['_id']
        
        # Convert datetime strings to datetime objects if needed
        if 'created_at' in project and isinstance(project['created_at'], str):
            try:
                project['created_at'] = datetime.fromisoformat(project['created_at'].replace('Z', '+00:00'))
            except ValueError:
                pass
        
        return ProjectRecord(**project)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get project: {e}")
        raise HTTPException(status_code=500, detail="Failed to get project")

@api_router.put("/projects/{project_code}", response_model=ProjectRecord)
async def update_project(project_code: str, project_update: ProjectUpdate, admin: str = Depends(authenticate_admin)):
    """Update an existing project"""
    try:
        # Get existing project
        existing_project = await db.projects.find_one({"project_code": project_code})
        if not existing_project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # Prepare update data
        update_data = project_update.dict(exclude_unset=True)
        
        # Validate dates if provided
        for date_field in ['start_date_iso', 'end_date_iso']:
            if date_field in update_data and update_data[date_field]:
                try:
                    datetime.strptime(update_data[date_field], '%Y-%m-%d')
                except ValueError:
                    raise HTTPException(status_code=400, detail=f"Invalid {date_field.replace('_', ' ')} format. Use YYYY-MM-DD")
        
        # Validate status if provided
        if "status" in update_data and update_data["status"]:
            valid_statuses = ['planned', 'active', 'completed', 'on-hold', 'cancelled']
            if update_data["status"] not in valid_statuses:
                raise HTTPException(status_code=400, detail=f"Status must be one of: {', '.join(valid_statuses)}")
        
        # Validate currency if provided
        if "budget_currency" in update_data and update_data["budget_currency"]:
            if update_data["budget_currency"] not in ['USD', 'LRD']:
                raise HTTPException(status_code=400, detail="Budget currency must be USD or LRD")
        
        # Validate budget amount if provided
        if "budget_amount" in update_data and update_data["budget_amount"] is not None:
            if update_data["budget_amount"] < 0:
                raise HTTPException(status_code=400, detail="Budget amount must be positive")
        
        # Update in database
        await db.projects.update_one(
            {"project_code": project_code},
            {"$set": update_data}
        )
        
        # Get updated project
        updated_project = await db.projects.find_one({"project_code": project_code})
        if '_id' in updated_project:
            del updated_project['_id']
        
        # Convert datetime strings to datetime objects if needed
        if 'created_at' in updated_project and isinstance(updated_project['created_at'], str):
            try:
                updated_project['created_at'] = datetime.fromisoformat(updated_project['created_at'].replace('Z', '+00:00'))
            except ValueError:
                pass
        
        return ProjectRecord(**updated_project)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update project: {e}")
        raise HTTPException(status_code=500, detail="Failed to update project")

@api_router.delete("/projects/{project_code}")
async def delete_project(project_code: str, admin: str = Depends(authenticate_admin)):
    """Delete a project"""
    try:
        result = await db.projects.delete_one({"project_code": project_code})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Project not found")
        return {"message": "Project deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete project: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete project")

@api_router.get("/projects/{project_code}/donations")
async def get_project_donations(project_code: str, admin: str = Depends(authenticate_admin)):
    """Get donations totaled by project"""
    try:
        # Aggregate donations by currency for this project
        pipeline = [
            {"$match": {"project_code": project_code}},
            {"$group": {
                "_id": "$amount_currency",
                "total": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }}
        ]
        
        donation_results = await db.donations.aggregate(pipeline).to_list(10)
        
        # Format results
        totals = {"USD": 0, "LRD": 0, "total_donations": 0}
        for result in donation_results:
            currency = result["_id"] or "USD"
            totals[currency] = result["total"]
            totals["total_donations"] += result["count"]
        
        # Get recent donations for this project
        recent_donations = await db.donations.find(
            {"project_code": project_code}
        ).sort("date_iso", -1).limit(5).to_list(5)
        
        # Clean up MongoDB _id fields
        for donation in recent_donations:
            if '_id' in donation:
                del donation['_id']
        
        return {
            "project_code": project_code,
            "totals": totals,
            "recent_donations": recent_donations
        }
        
    except Exception as e:
        logger.error(f"Failed to get project donations: {e}")
        raise HTTPException(status_code=500, detail="Failed to get project donations")

@api_router.get("/projects/{project_code}/stories")
async def get_project_stories(project_code: str, admin: str = Depends(authenticate_admin)):
    """Get recent stories linked to this project"""
    try:
        # For now, we'll link stories by matching project code in story text or location
        # This is a simple implementation - in a real system you'd have explicit project links
        recent_stories = await db.stories.find({
            "$or": [
                {"story_text": {"$regex": project_code, "$options": "i"}},
                {"location": {"$regex": project_code, "$options": "i"}}
            ]
        }).sort("date_iso", -1).limit(5).to_list(5)
        
        # Clean up MongoDB _id fields
        for story in recent_stories:
            if '_id' in story:
                del story['_id']
        
        return {
            "project_code": project_code,
            "recent_stories": recent_stories,
            "total_stories": len(recent_stories)
        }
        
    except Exception as e:
        logger.error(f"Failed to get project stories: {e}")
        raise HTTPException(status_code=500, detail="Failed to get project stories")

@api_router.get("/projects/export/csv")
async def export_projects_csv(
    status: Optional[str] = None,
    country: Optional[str] = None,
    manager: Optional[str] = None,
    admin: str = Depends(authenticate_admin)
):
    """Export projects as CSV"""
    try:
        # Build filter
        filter_dict = {}
        if status:
            filter_dict["status"] = status
        if country:
            filter_dict["country"] = country
        if manager:
            filter_dict["manager"] = manager
        
        # Get all matching projects
        projects = await db.projects.find(filter_dict).sort("created_at", -1).to_list(10000)
        
        # Create CSV content
        csv_content = "project_code,name,description_short,start_date_iso,end_date_iso,status,budget_currency,budget_amount,manager,country,tags\n"
        
        for project in projects:
            row = [
                project.get("project_code", ""),
                project.get("name", ""),
                project.get("description_short", "").replace('"', '""'),  # Escape quotes
                project.get("start_date_iso", ""),
                project.get("end_date_iso", ""),
                project.get("status", ""),
                project.get("budget_currency", ""),
                str(project.get("budget_amount", "")),
                project.get("manager", ""),
                project.get("country", ""),
                project.get("tags", "")
            ]
            csv_content += '"' + '","'.join(row) + '"\n'
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"projects_export_{timestamp}.csv"
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export projects CSV: {e}")
        raise HTTPException(status_code=500, detail="Failed to export projects CSV")

@api_router.get("/projects/export/xlsx")
async def export_projects_xlsx(
    status: Optional[str] = None,
    country: Optional[str] = None,
    manager: Optional[str] = None,
    admin: str = Depends(authenticate_admin)
):
    """Export projects as XLSX"""
    try:
        # Build filter
        filter_dict = {}
        if status:
            filter_dict["status"] = status
        if country:
            filter_dict["country"] = country
        if manager:
            filter_dict["manager"] = manager
        
        # Get all matching projects
        projects = await db.projects.find(filter_dict).sort("created_at", -1).to_list(10000)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        
        # Headers
        headers = ["Project Code", "Name", "Description", "Start Date", "End Date", "Status", "Budget Currency", "Budget Amount", "Manager", "Country", "Tags"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for project in projects:
            row = [
                project.get("project_code", ""),
                project.get("name", ""),
                project.get("description_short", ""),
                project.get("start_date_iso", ""),
                project.get("end_date_iso", ""),
                project.get("status", ""),
                project.get("budget_currency", ""),
                project.get("budget_amount", ""),
                project.get("manager", ""),
                project.get("country", ""),
                project.get("tags", "")
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"projects_export_{timestamp}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export projects XLSX: {e}")
        raise HTTPException(status_code=500, detail="Failed to export projects XLSX")

@api_router.get("/projects/filter-stats")
async def get_projects_filter_stats(admin: str = Depends(authenticate_admin)):
    """Get projects statistics for filters"""
    try:
        # Get unique statuses
        statuses = await db.projects.distinct("status")
        statuses = [s for s in statuses if s]  # Remove empty/null values
        
        # Get unique countries
        countries = await db.projects.distinct("country")
        countries = [c for c in countries if c]  # Remove empty/null values
        
        # Get unique managers
        managers = await db.projects.distinct("manager")
        managers = [m for m in managers if m]  # Remove empty/null values
        
        return {
            "statuses": sorted(statuses),
            "countries": sorted(countries),
            "managers": sorted(managers),
            "total_projects": await db.projects.count_documents({})
        }
        
    except Exception as e:
        logger.error(f"Failed to get projects filter stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get projects filter stats")

# Dashboard Models
class DashboardStats(BaseModel):
    visitors_this_month: int = 0
    donations_this_month: float = 0.0
    income_this_month: float = 0.0
    expenses_this_month: float = 0.0
    open_reminders: int = 0
    approved_stories: int = 0
    last_updated: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DonationByProject(BaseModel):
    project_name: str
    amount: float
    percentage: float

class IncomeExpenseData(BaseModel):
    month: str
    income: float
    expenses: float

# Dashboard endpoints with authentication
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(admin: str = Depends(authenticate_admin)):
    """Get dashboard statistics"""
    try:
        # Get current month start and end
        now = datetime.now(timezone.utc)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = month_start.replace(month=month_start.month + 1) if month_start.month < 12 else month_start.replace(year=month_start.year + 1, month=1)
        
        # Visitors this month (from visitor analytics)
        visitors_this_month = await db.visitor_analytics.count_documents({
            "timestamp": {
                "$gte": month_start.isoformat(),
                "$lt": next_month.isoformat()
            }
        })
        
        # Donations this month (sum from donations collection)
        donations_pipeline = [
            {
                "$match": {
                    "created_at": {
                        "$gte": month_start.isoformat(),
                        "$lt": next_month.isoformat()
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total": {"$sum": "$amount"}
                }
            }
        ]
        donations_result = await db.donations.aggregate(donations_pipeline).to_list(1)
        donations_this_month = donations_result[0]["total"] if donations_result else 0.0
        
        # Major gift pledges this month
        pledges_pipeline = [
            {
                "$match": {
                    "created_at": {
                        "$gte": month_start.isoformat(),
                        "$lt": next_month.isoformat()
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total": {"$sum": "$amount"}
                }
            }
        ]
        pledges_result = await db.major_gift_pledges.aggregate(pledges_pipeline).to_list(1)
        pledges_this_month = pledges_result[0]["total"] if pledges_result else 0.0
        
        # Total donations including pledges
        total_donations = donations_this_month + pledges_this_month
        
        # Mock data for income and expenses (in real app, would come from accounting system)
        income_this_month = total_donations + 5000  # Add regular donations/support
        expenses_this_month = 3200  # Operating expenses
        
        # Open reminders (mock data - in real app would come from reminders collection)
        open_reminders = 7
        
        # Approved stories (from impact stories)
        approved_stories = await db.impact_stories.count_documents({
            "is_featured": True
        })
        
        return DashboardStats(
            visitors_this_month=visitors_this_month,
            donations_this_month=total_donations,
            income_this_month=income_this_month,
            expenses_this_month=expenses_this_month,
            open_reminders=open_reminders,
            approved_stories=approved_stories,
            last_updated=datetime.now(timezone.utc)
        )
        
    except Exception as e:
        logger.error(f"Failed to get dashboard stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get dashboard stats")

@api_router.get("/dashboard/donations-by-project")
async def get_donations_by_project(admin: str = Depends(authenticate_admin)):
    """Get donations breakdown by project for pie chart"""
    try:
        # Get current month start and end
        now = datetime.now(timezone.utc)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = month_start.replace(month=month_start.month + 1) if month_start.month < 12 else month_start.replace(year=month_start.year + 1, month=1)
        
        # Get major gift pledges by designation
        pledges_pipeline = [
            {
                "$match": {
                    "created_at": {
                        "$gte": month_start.isoformat(),
                        "$lt": next_month.isoformat()
                    }
                }
            },
            {
                "$group": {
                    "_id": "$designation",
                    "total": {"$sum": "$amount"}
                }
            }
        ]
        pledges_by_project = await db.major_gift_pledges.aggregate(pledges_pipeline).to_list(10)
        
        # Get general donations
        general_donations_pipeline = [
            {
                "$match": {
                    "created_at": {
                        "$gte": month_start.isoformat(),
                        "$lt": next_month.isoformat()
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total": {"$sum": "$amount"}
                }
            }
        ]
        general_result = await db.donations.aggregate(general_donations_pipeline).to_list(1)
        general_amount = general_result[0]["total"] if general_result else 0.0
        
        # Combine all donations
        project_donations = []
        total_amount = general_amount
        
        # Add general donations
        if general_amount > 0:
            project_donations.append({
                "project_name": "General Support",
                "amount": general_amount
            })
        
        # Add major gift pledges by project
        for pledge in pledges_by_project:
            project_name = pledge["_id"] or "Unspecified"
            amount = pledge["total"]
            total_amount += amount
            project_donations.append({
                "project_name": project_name,
                "amount": amount
            })
        
        # Add mock data if no real donations
        if total_amount == 0:
            project_donations = [
                {"project_name": "Solar Array", "amount": 1200.0},
                {"project_name": "Studio Equipment", "amount": 800.0},
                {"project_name": "General Support", "amount": 500.0},
                {"project_name": "Transmitter", "amount": 300.0}
            ]
            total_amount = sum(d["amount"] for d in project_donations)
        
        # Calculate percentages
        result = []
        for donation in project_donations:
            percentage = (donation["amount"] / total_amount * 100) if total_amount > 0 else 0
            result.append(DonationByProject(
                project_name=donation["project_name"],
                amount=donation["amount"],
                percentage=round(percentage, 1)
            ))
        
        return result
        
    except Exception as e:
        logger.error(f"Failed to get donations by project: {e}")
        raise HTTPException(status_code=500, detail="Failed to get donations by project")

@api_router.get("/dashboard/income-expenses")
async def get_income_expenses(admin: str = Depends(authenticate_admin)):
    """Get income vs expenses data for bar chart"""
    try:
        # Get current month start and end
        now = datetime.now(timezone.utc)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = month_start.replace(month=month_start.month + 1) if month_start.month < 12 else month_start.replace(year=month_start.year + 1, month=1)
        
        # Get donations for this month
        donations_pipeline = [
            {
                "$match": {
                    "created_at": {
                        "$gte": month_start.isoformat(),
                        "$lt": next_month.isoformat()
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total": {"$sum": "$amount"}
                }
            }
        ]
        donations_result = await db.donations.aggregate(donations_pipeline).to_list(1)
        donations_this_month = donations_result[0]["total"] if donations_result else 0.0
        
        # Get pledges for this month
        pledges_pipeline = [
            {
                "$match": {
                    "created_at": {
                        "$gte": month_start.isoformat(),
                        "$lt": next_month.isoformat()
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total": {"$sum": "$amount"}
                }
            }
        ]
        pledges_result = await db.major_gift_pledges.aggregate(pledges_pipeline).to_list(1)
        pledges_this_month = pledges_result[0]["total"] if pledges_result else 0.0
        
        # Calculate income and expenses
        total_donations = donations_this_month + pledges_this_month
        income = total_donations + 5000  # Add regular support/grants
        expenses = 3200  # Operating expenses (utilities, maintenance, salaries, etc.)
        
        # Add mock data if no real data
        if income == 5000:  # Only base amount, no donations
            income = 8500
            expenses = 3200
        
        return IncomeExpenseData(
            month=now.strftime("%B %Y"),
            income=income,
            expenses=expenses
        )
        
    except Exception as e:
        logger.error(f"Failed to get income expenses data: {e}")
        raise HTTPException(status_code=500, detail="Failed to get income expenses data")

# =============================================================================
# AI PROGRAM ASSISTANT ENDPOINTS
# =============================================================================

# Import AI integration libraries
from emergentintegrations.llm.chat import LlmChat, UserMessage

# Pydantic models for AI Program Assistant
class ProgramContent(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    content: str  # Original program content/transcript
    summary: Optional[str] = None  # AI-generated summary
    highlights: Optional[List[str]] = None  # AI-generated highlights
    keywords: Optional[List[str]] = None  # AI-extracted keywords
    language: str = "en"  # Original language
    translated_content: Optional[Dict[str, str]] = None  # Translations {"fr": "content"}
    date_aired: Optional[str] = None  # YYYY-MM-DD format
    duration_minutes: Optional[int] = None
    presenter: Optional[str] = None
    program_type: Optional[str] = None  # e.g., "devotion", "music", "news"
    audio_url: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ProgramCreate(BaseModel):
    title: str
    description: Optional[str] = None
    content: str
    language: str = "en"
    date_aired: Optional[str] = None
    duration_minutes: Optional[int] = None
    presenter: Optional[str] = None
    program_type: Optional[str] = None
    audio_url: Optional[str] = None

class AIAnalysisRequest(BaseModel):
    program_id: str
    analysis_type: str = "summary"  # "summary", "highlights", "keywords", "translate"
    target_language: Optional[str] = "fr"  # For translation requests

class AIAnalysisResponse(BaseModel):
    program_id: str
    analysis_type: str
    result: Dict[str, Any]
    processing_time: float
    model_used: str

class ProgramSearchRequest(BaseModel):
    query: str
    language: Optional[str] = None
    program_type: Optional[str] = None
    date_range: Optional[Dict[str, str]] = None  # {"start": "YYYY-MM-DD", "end": "YYYY-MM-DD"}
    limit: int = 10

# AI Helper Functions
async def get_ai_client(model_type: str = "openai"):
    """Initialize AI client with Emergent LLM key"""
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        raise HTTPException(status_code=500, detail="AI service not configured")
    
    session_id = f"kioo_program_assistant_{datetime.now().timestamp()}"
    
    if model_type == "openai":
        chat = LlmChat(
            api_key=api_key,
            session_id=session_id,
            system_message="You are an AI assistant for Kioo Radio. Help analyze, summarize, and enhance radio program content. Focus on faith-based content, community impact, and spiritual growth themes."
        ).with_model("openai", "gpt-4o")
    elif model_type == "claude":
        chat = LlmChat(
            api_key=api_key,
            session_id=session_id,
            system_message="You are Claude, an AI assistant for Kioo Radio. Help analyze, summarize, and enhance faith-based radio program content with accuracy and cultural sensitivity."
        ).with_model("anthropic", "claude-3-7-sonnet-20250219")
    else:
        # Default to OpenAI
        chat = LlmChat(
            api_key=api_key,
            session_id=session_id,
            system_message="You are an AI assistant for Kioo Radio. Help analyze, summarize, and enhance radio program content."
        ).with_model("openai", "gpt-4o-mini")
    
    return chat

async def ai_summarize_content(content: str, language: str = "en") -> str:
    """Generate AI summary of program content"""
    try:
        chat = await get_ai_client("openai")
        
        prompt = f"""
        Please create a concise summary of this radio program content in {language}:
        
        Content: {content[:4000]}  # Limit content for token efficiency
        
        Focus on:
        - Main themes and messages
        - Key spiritual insights
        - Community impact mentioned
        - Practical takeaways for listeners
        
        Keep the summary under 200 words and maintain the spiritual tone appropriate for Kioo Radio's faith-based mission.
        """
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        return response
        
    except Exception as e:
        logger.error(f"AI summarization failed: {e}")
        return "Summary generation failed. Please try again later."

async def ai_extract_highlights(content: str, language: str = "en") -> List[str]:
    """Extract key highlights from program content"""
    try:
        chat = await get_ai_client("claude")
        
        prompt = f"""
        Extract 5-7 key highlights from this radio program content in {language}:
        
        Content: {content[:4000]}
        
        Return highlights as a JSON array of strings. Focus on:
        - Inspirational quotes
        - Key spiritual messages
        - Community announcements
        - Prayer requests or testimonies
        - Practical spiritual advice
        
        Example format: ["Highlight 1", "Highlight 2", "Highlight 3"]
        """
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        # Try to parse JSON response
        import json
        try:
            highlights = json.loads(response)
            if isinstance(highlights, list):
                return highlights[:7]  # Limit to 7 highlights
        except:
            # If JSON parsing fails, split by lines
            lines = response.strip().split('\n')
            return [line.strip('- ').strip() for line in lines if line.strip()][:7]
            
    except Exception as e:
        logger.error(f"AI highlight extraction failed: {e}")
        return ["Content analysis unavailable"]

async def ai_extract_keywords(content: str, language: str = "en") -> List[str]:
    """Extract relevant keywords from program content"""
    try:
        chat = await get_ai_client("openai")
        
        prompt = f"""
        Extract 10-15 relevant keywords from this radio program content in {language}:
        
        Content: {content[:4000]}
        
        Focus on:
        - Spiritual themes
        - Biblical references
        - Community topics
        - Faith concepts
        - Practical life topics
        
        Return as comma-separated keywords only, no explanations.
        """
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        # Parse keywords
        keywords = [kw.strip() for kw in response.split(',') if kw.strip()]
        return keywords[:15]  # Limit to 15 keywords
        
    except Exception as e:
        logger.error(f"AI keyword extraction failed: {e}")
        return ["faith", "community", "spiritual growth"]

async def ai_translate_content(content: str, target_language: str = "fr") -> str:
    """Translate program content to target language"""
    try:
        chat = await get_ai_client("claude")
        
        lang_names = {"fr": "French", "en": "English"}
        target_lang_name = lang_names.get(target_language, target_language)
        
        prompt = f"""
        Translate the following radio program content to {target_lang_name}. 
        Maintain the spiritual tone and cultural context appropriate for African French-speaking audiences:
        
        Content: {content[:3000]}
        
        Provide only the translation, no additional commentary.
        """
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        return response
        
    except Exception as e:
        logger.error(f"AI translation failed: {e}")
        return f"Translation to {target_language} unavailable"

# AI Program Assistant API Endpoints

@api_router.post("/ai-programs", response_model=ProgramContent)
async def create_ai_program(
    program: ProgramCreate,
    admin: str = Depends(authenticate_admin)
):
    """Create a new AI program with AI analysis"""
    try:
        # Create program document
        program_doc = ProgramContent(**program.dict())
        
        # Store in database (use separate collection for AI programs)
        result = await db.ai_programs.insert_one(program_doc.dict())
        program_doc.id = str(result.inserted_id) if hasattr(result, 'inserted_id') else program_doc.id
        
        # Start AI analysis in background (non-blocking for user experience)
        try:
            # Generate summary
            program_doc.summary = await ai_summarize_content(program.content, program.language)
            
            # Extract highlights
            program_doc.highlights = await ai_extract_highlights(program.content, program.language)
            
            # Extract keywords
            program_doc.keywords = await ai_extract_keywords(program.content, program.language)
            
            # Update document with AI analysis
            program_doc.updated_at = datetime.now(timezone.utc).isoformat()
            await db.ai_programs.update_one(
                {"id": program_doc.id},
                {"$set": program_doc.dict()}
            )
            
        except Exception as ai_error:
            logger.warning(f"AI analysis failed for program {program_doc.id}: {ai_error}")
        
        return program_doc
        
    except Exception as e:
        logger.error(f"Failed to create AI program: {e}")
        raise HTTPException(status_code=500, detail="Failed to create AI program")

@api_router.get("/ai-programs", response_model=List[ProgramContent])
async def get_ai_programs(
    limit: int = 20,
    skip: int = 0,
    program_type: Optional[str] = None,
    language: Optional[str] = None,
    presenter: Optional[str] = None,
    admin: str = Depends(authenticate_admin)
):
    """Get all AI programs with optional filtering"""
    try:
        # Build filter
        filter_dict = {}
        if program_type:
            filter_dict["program_type"] = program_type
        if language:
            filter_dict["language"] = language
        if presenter:
            filter_dict["presenter"] = {"$regex": presenter, "$options": "i"}
        
        # Get programs from AI programs collection
        programs = await db.ai_programs.find(filter_dict).sort("date_aired", -1).skip(skip).limit(limit).to_list(limit)
        
        return [ProgramContent(**program) for program in programs]
        
    except Exception as e:
        logger.error(f"Failed to get AI programs: {e}")
        raise HTTPException(status_code=500, detail="Failed to get AI programs")

@api_router.get("/ai-programs/{program_id}", response_model=ProgramContent)
async def get_ai_program(
    program_id: str,
    admin: str = Depends(authenticate_admin)
):
    """Get specific AI program by ID"""
    try:
        program = await db.ai_programs.find_one({"id": program_id})
        if not program:
            raise HTTPException(status_code=404, detail="AI Program not found")
        
        return ProgramContent(**program)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get AI program: {e}")
        raise HTTPException(status_code=500, detail="Failed to get AI program")

@api_router.post("/ai-programs/{program_id}/analyze", response_model=AIAnalysisResponse)
async def analyze_ai_program(
    program_id: str,
    request: AIAnalysisRequest,
    admin: str = Depends(authenticate_admin)
):
    """Trigger AI analysis for a specific AI program"""
    try:
        start_time = datetime.now()
        
        # Get program from AI programs collection
        program = await db.ai_programs.find_one({"id": program_id})
        if not program:
            raise HTTPException(status_code=404, detail="AI Program not found")
        
        result = {}
        model_used = "gpt-4o"
        
        if request.analysis_type == "summary":
            result["summary"] = await ai_summarize_content(program["content"], program.get("language", "en"))
            await db.ai_programs.update_one(
                {"id": program_id},
                {"$set": {"summary": result["summary"], "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            
        elif request.analysis_type == "highlights":
            result["highlights"] = await ai_extract_highlights(program["content"], program.get("language", "en"))
            await db.ai_programs.update_one(
                {"id": program_id},
                {"$set": {"highlights": result["highlights"], "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            
        elif request.analysis_type == "keywords":
            result["keywords"] = await ai_extract_keywords(program["content"], program.get("language", "en"))
            await db.ai_programs.update_one(
                {"id": program_id},
                {"$set": {"keywords": result["keywords"], "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            
        elif request.analysis_type == "translate":
            target_lang = request.target_language or "fr"
            translation = await ai_translate_content(program["content"], target_lang)
            
            # Update translated content
            translated_content = program.get("translated_content", {})
            translated_content[target_lang] = translation
            
            result["translation"] = translation
            result["target_language"] = target_lang
            
            await db.ai_programs.update_one(
                {"id": program_id},
                {"$set": {"translated_content": translated_content, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            model_used = "claude-3-7-sonnet"
            
        else:
            raise HTTPException(status_code=400, detail="Invalid analysis type")
        
        processing_time = (datetime.now() - start_time).total_seconds()
        
        return AIAnalysisResponse(
            program_id=program_id,
            analysis_type=request.analysis_type,
            result=result,
            processing_time=processing_time,
            model_used=model_used
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to analyze AI program: {e}")
        raise HTTPException(status_code=500, detail="Failed to analyze AI program")

@api_router.post("/ai-programs/search")
async def search_ai_programs(
    request: ProgramSearchRequest,
    admin: str = Depends(authenticate_admin)
):
    """AI-powered semantic search of AI programs"""
    try:
        # Build MongoDB text search query on AI programs collection
        search_filter = {"$text": {"$search": request.query}}
        
        # Additional filters
        if request.language:
            search_filter["language"] = request.language
        if request.program_type:
            search_filter["program_type"] = request.program_type
        if request.date_range:
            date_filter = {}
            if request.date_range.get("start"):
                date_filter["$gte"] = request.date_range["start"]
            if request.date_range.get("end"):
                date_filter["$lte"] = request.date_range["end"]
            if date_filter:
                search_filter["date_aired"] = date_filter
        
        # Perform search on AI programs collection
        programs = await db.ai_programs.find(
            search_filter,
            {"score": {"$meta": "textScore"}}
        ).sort([("score", {"$meta": "textScore"})]).limit(request.limit).to_list(request.limit)
        
        # If no text search results, fall back to keyword search
        if not programs:
            keyword_filter = {
                "$or": [
                    {"title": {"$regex": request.query, "$options": "i"}},
                    {"description": {"$regex": request.query, "$options": "i"}},
                    {"content": {"$regex": request.query, "$options": "i"}},
                    {"keywords": {"$in": [request.query]}}
                ]
            }
            
            # Add additional filters
            if request.language:
                keyword_filter["language"] = request.language
            if request.program_type:
                keyword_filter["program_type"] = request.program_type
            if request.date_range:
                date_filter = {}
                if request.date_range.get("start"):
                    date_filter["$gte"] = request.date_range["start"]
                if request.date_range.get("end"):
                    date_filter["$lte"] = request.date_range["end"]
                if date_filter:
                    keyword_filter["date_aired"] = date_filter
            
            programs = await db.ai_programs.find(keyword_filter).sort("date_aired", -1).limit(request.limit).to_list(request.limit)
        
        return {
            "query": request.query,
            "total_results": len(programs),
            "programs": [ProgramContent(**program) for program in programs]
        }
        
    except Exception as e:
        logger.error(f"AI Program search failed: {e}")
        raise HTTPException(status_code=500, detail="AI Program search failed")

@api_router.get("/ai-programs/stats/overview")
async def get_ai_program_stats(admin: str = Depends(authenticate_admin)):
    """Get AI program statistics overview"""
    try:
        # Get basic counts from AI programs collection
        total_programs = await db.ai_programs.count_documents({})
        
        # Get counts by language
        language_pipeline = [
            {"$group": {"_id": "$language", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        language_stats = await db.ai_programs.aggregate(language_pipeline).to_list(10)
        
        # Get counts by program type
        type_pipeline = [
            {"$group": {"_id": "$program_type", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        type_stats = await db.ai_programs.aggregate(type_pipeline).to_list(10)
        
        # Get recent activity
        recent_programs = await db.ai_programs.find({}).sort("created_at", -1).limit(5).to_list(5)
        
        # Calculate AI analysis coverage
        programs_with_summary = await db.ai_programs.count_documents({"summary": {"$exists": True, "$ne": None}})
        programs_with_highlights = await db.ai_programs.count_documents({"highlights": {"$exists": True, "$ne": []}})
        programs_with_keywords = await db.ai_programs.count_documents({"keywords": {"$exists": True, "$ne": []}})
        
        return {
            "total_programs": total_programs,
            "ai_analysis_coverage": {
                "with_summary": programs_with_summary,
                "with_highlights": programs_with_highlights,
                "with_keywords": programs_with_keywords,
                "summary_percentage": round((programs_with_summary / max(total_programs, 1)) * 100, 1),
                "highlights_percentage": round((programs_with_highlights / max(total_programs, 1)) * 100, 1),
                "keywords_percentage": round((programs_with_keywords / max(total_programs, 1)) * 100, 1)
            },
            "by_language": {item["_id"] or "unknown": item["count"] for item in language_stats},
            "by_type": {item["_id"] or "unknown": item["count"] for item in type_stats},
            "recent_programs": [
                {
                    "id": prog["id"],
                    "title": prog.get("title", "Untitled"),
                    "date_aired": prog.get("date_aired"),
                    "language": prog.get("language", "unknown"),
                    "created_at": prog.get("created_at")
                }
                for prog in recent_programs
            ]
        }
        
    except Exception as e:
        logger.error(f"Failed to get AI program stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get AI program stats")

# Include the router in the main app
app.include_router(api_router)

# Configure logging for production
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# Startup event
@app.on_event("startup")
async def startup_db_client():
    """Initialize database connection on startup"""
    try:
        # Test the connection
        await client.admin.command('ping')
        logger.info("Successfully connected to MongoDB")
        
        # Create text indexes for program search
        try:
            await db.programs.create_index([
                ("title", "text"),
                ("description", "text"),
                ("content", "text"),
                ("keywords", "text")
            ])
            logger.info("Created text search indexes for programs collection")
        except Exception as index_error:
            logger.warning(f"Failed to create search indexes: {index_error}")
            
    except Exception as e:
        logger.error(f"Failed to connect to MongoDB: {e}")
        raise

# Shutdown event
@app.on_event("shutdown")
async def shutdown_db_client():
    """Close database connection on shutdown"""
    try:
        client.close()
        logger.info("MongoDB connection closed")
    except Exception as e:
        logger.error(f"Error closing MongoDB connection: {e}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)